"""
Table View Service (Simplified for Actual Schema)
Handles data retrieval and filtering for the main TableView page.
Provides paginated, searchable, and filterable complaints/incidents list.

NOTE: This version works with the actual APP_IncidentCase schema which has:
- IncidentRequestCaseID (primary key)
- ComplaintText, ImmediateAction, TakenAction
- FeedbackRecievedDate, PatientName, DoctorName, DoctorID
- IssuingOrgUnitID (organizational unit)
- DomainID, CategoryID, SubCategoryID, ClassificationID
- SeverityID, StageID, HarmLevelID, CaseStatusID
- CreatedAt, CreatedByUserID, InOut, ClinicalRiskTypeID, FeedbackIntentTypeID, BuildingID
"""

from typing import Dict, List, Optional, Any, Literal
from datetime import datetime, date
import pyodbc
from dateutil.relativedelta import relativedelta
from io import BytesIO
from core.database import get_connection
from core.table_config import HR_EMPLOYEES_TABLE, DOCTORS_TABLE
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from ..db_layer.database import get_connection
from ..db_layer.satisfaction_db import get_satisfactions_by_cases


# ==================== HELPER FUNCTIONS ====================

def _parse_multi_int(value: Optional[str]) -> Optional[List[int]]:
    """
    Parse a comma-separated string of IDs (e.g. "5,12,7") into a list of ints.
    Also accepts a single bare int/str for backward compatibility with callers
    that still pass one scalar value. Returns None if nothing usable is present.
    """
    if value is None or value == "":
        return None
    try:
        ids = [int(v.strip()) for v in str(value).split(",") if v.strip() != ""]
    except ValueError:
        return None
    return ids or None


def _in_clause(column_expr: str, ids: List[int]) -> str:
    """Build a parameterized SQL IN (...) clause for the given column expression."""
    placeholders = ",".join("?" * len(ids))
    return f"{column_expr} IN ({placeholders})"

def _get_workflow_status(incident_id: int, cursor=None) -> Optional[Dict[str, Any]]:
    """
    Get workflow status for an incident (subcases information).
    
    Returns:
        Dictionary with workflow status info or None if no subcases
        {
            "has_subcases": bool,
            "open_subcase_count": int,
            "force_closed": bool,
            "subcases": [
                {
                    "subcase_id": int,
                    "status": str,
                    "target_org_unit": str,
                    "target_org_unit_id": int
                },
                ...
            ]
        }
    """
    close_cursor = False
    if cursor is None:
        conn = get_connection()
        cursor = conn.cursor()
        close_cursor = True
    
    try:
        query = """
            SELECT 
                s.SubcaseID,
                s.Status,
                s.TargetOrgUnitID,
                org.Name as TargetOrgUnitName
            FROM dbo.APP_AdministrativeSubcase s
            LEFT JOIN AdminsrationUnit org ON s.TargetOrgUnitID = org.UniqueID
            WHERE s.IncidentRequestCaseID = ?
            ORDER BY s.CreatedAt ASC
        """
        
        cursor.execute(query, (incident_id,))
        rows = cursor.fetchall()
        
        if not rows:
            return None
        
        subcases = []
        open_count = 0
        force_closed_count = 0
        
        for row in rows:
            status = row.Status
            if status == 'FORCE_CLOSED':
                force_closed_count += 1
            elif status not in ['ADMIN_APPROVED', 'SECTION_DENIED']:
                open_count += 1
            
            subcases.append({
                "subcase_id": row.SubcaseID,
                "status": status,
                "target_org_unit": row.TargetOrgUnitName,
                "target_org_unit_id": row.TargetOrgUnitID
            })
        
        return {
            "has_subcases": True,
            "open_subcase_count": open_count,
            "force_closed": force_closed_count > 0,
            "subcases": subcases
        }
    
    finally:
        if close_cursor:
            cursor.close()
            conn.close()


def _get_workflow_statuses_batch(case_ids: list, cursor) -> dict:
    """
    Fetch workflow statuses for a list of case IDs in a single query.
    Returns dict keyed by IncidentRequestCaseID.
    Includes is_late derived from per-level LateReply flags.
    """
    if not case_ids:
        return {}

    placeholders = ",".join("?" * len(case_ids))
    query = f"""
        SELECT
            s.IncidentRequestCaseID,
            s.SubcaseID,
            s.Status,
            s.TargetOrgUnitID,
            org.Name AS TargetOrgUnitName,
            s.SectionLateReply,
            s.DepartmentLateReply,
            s.AdministrationLateReply
        FROM dbo.APP_AdministrativeSubcase s
        LEFT JOIN AdminsrationUnit org ON s.TargetOrgUnitID = org.UniqueID
        WHERE s.IncidentRequestCaseID IN ({placeholders})
        ORDER BY s.IncidentRequestCaseID, s.CreatedAt ASC
    """
    cursor.execute(query, case_ids)
    rows = cursor.fetchall()

    result = {}
    for row in rows:
        cid = row.IncidentRequestCaseID
        if cid not in result:
            result[cid] = {
                "has_subcases": True,
                "open_subcase_count": 0,
                "force_closed": False,
                "is_late": False,
                "subcases": [],
            }
        status = row.Status
        if status == "FORCE_CLOSED":
            result[cid]["force_closed"] = True
        elif status not in ("ADMIN_APPROVED", "SECTION_DENIED"):
            result[cid]["open_subcase_count"] += 1

        if row.SectionLateReply or row.DepartmentLateReply or row.AdministrationLateReply:
            result[cid]["is_late"] = True

        result[cid]["subcases"].append({
            "subcase_id": row.SubcaseID,
            "status": status,
            "target_org_unit": row.TargetOrgUnitName,
            "target_org_unit_id": row.TargetOrgUnitID,
        })

    return result


# ==================== MAIN ENDPOINTS ====================

def get_complaints_paginated(
    search: Optional[str] = None,
    issuing_org_unit_id: Optional[int] = None,
    target_department_id: Optional[int] = None,
    target_dept_parent_id: Optional[int] = None,
    target_admin_id: Optional[int] = None,
    domain_id: Optional[str] = None,
    category_id: Optional[str] = None,
    subcategory_id: Optional[str] = None,
    classification_id: Optional[str] = None,
    severity_id: Optional[str] = None,
    stage_id: Optional[str] = None,
    harm_level_id: Optional[str] = None,
    case_status_id: Optional[str] = None,
    clinical_risk_type_id: Optional[str] = None,
    feedback_intent_type_id: Optional[str] = None,
    source_id: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    sort_by: str = "FeedbackRecievedDate",
    sort_order: str = "desc",
    page: int = 1,
    page_size: int = 50,
    view: str = "complete",
    tab: Optional[str] = None,  # 'preparation' | 'workflow' | None
    restrict_to_unit_ids: Optional[set] = None,  # org-scoped viewer enforcement
) -> Dict[str, Any]:
    """
    Fetch paginated complaints with search and filtering.
    
    Returns:
        Dictionary with complaints array, pagination info, and applied filters.
    """
    # Validate pagination
    if page < 1:
        raise ValueError("Page must be >= 1")
    if page_size < 1 or page_size > 500:
        raise ValueError("Page size must be between 1 and 500")
    
    # Sort field whitelist — maps a stable sort key to a literal, trusted SQL expression.
    # sort_by (user input) is only ever used as a dict key lookup, never interpolated
    # directly into SQL, so no injection surface exists here.
    # SeverityID/HarmSeverityOrder use the lookup tables' rank columns (SeverityOrder),
    # not the raw FK id, since FK assignment order is not guaranteed to match severity rank.
    SORT_FIELD_MAP = {
        "FeedbackRecievedDate": "c.FeedbackRecievedDate",
        "IncidentDate": "c.IncidentDate",
        "IncidentRequestCaseID": "CAST(c.IncidentRequestCaseID AS INT)",
        "id": "CAST(c.IncidentRequestCaseID AS INT)",
        "CreatedAt": "c.CreatedAt",
        "UpdatedAt": "c.UpdatedAt",
        "PublicationDate": "subcase_pub.CreatedAt",
        "PatientName": "c.PatientName",
        "SeverityID": "severity.SeverityOrder",
        "HarmSeverityOrder": "harm.SeverityOrder",
        "IncidentNumber": "i.incident_number",
        "SourceName": "source.SourceName",
        "FeedbackIntentTypeName": "feedback_intent.NameEn",
        "ClinicalRiskTypeName": "clinical_risk.Name",
        "StageName": "stage.StageName",
        "DomainName": "domain.DomainName",
        "CategoryName": "category.CategoryName",
        "SubCategoryName": "subcategory.SubCategoryName",
        "ClassificationName": "classification.Classification_AR",
        "StatusDisplayOrder": "status.DisplayOrder",
        "TargetDepartmentName": "target_depts.target_dept_names",
        "RcaReplies": "rca_replies.rca_replies_text",
        "ComplaintSummary": "LEFT(c.ComplaintText, 150)",
        "CustomerServiceDecision": "subcase_ans.PatientServicesDecisionText",
        "CustomerServiceDecisionDate": "subcase_ans.PatientServicesDecisionAt",
        "SectionEntry": "subcase_ans.SectionEntryTimestamp",
        "SectionDeadline": "subcase_ans.SectionDeadlineAt",
        "DepartmentEntry": "subcase_ans.DepartmentEntryTimestamp",
        "DepartmentDeadline": "subcase_ans.DepartmentDeadlineAt",
        "AdministrationEntry": "subcase_ans.AdministrationEntryTimestamp",
        "AdministrationDeadline": "subcase_ans.AdministrationDeadlineAt",
    }
    if sort_by not in SORT_FIELD_MAP:
        sort_by = "FeedbackRecievedDate"

    if sort_order.lower() not in ["asc", "desc"]:
        sort_order = "desc"

    # Build WHERE clause dynamically
    where_conditions = []
    params = []
    
    # Free-text search
    if search:
        search = search.strip()
    if search:
        search_condition = """(
            CAST(c.IncidentRequestCaseID AS VARCHAR) LIKE ?
            OR c.PatientName LIKE ?
            OR c.ComplaintText LIKE ?
            OR i.incident_number LIKE ?
        )"""
        where_conditions.append(search_condition)
        search_param = f"%{search}%"
        params.extend([search_param, search_param, search_param, search_param])
    
    # Organizational unit filter
    if issuing_org_unit_id:
        where_conditions.append("c.IssuingOrgUnitID = ?")
        params.append(issuing_org_unit_id)

    # Role-based scope enforcement: org viewer roles see only cases TARGETED AT their unit(s).
    # Uses APP_IncidentCaseTargetDepartment so a SECTION_ADMIN sees every complaint
    # that concerns their section, not just complaints that originated from it.
    if restrict_to_unit_ids is not None:
        if not restrict_to_unit_ids:
            # Empty set means no accessible units — return nothing
            where_conditions.append("1=0")
        else:
            placeholders = ",".join("?" * len(restrict_to_unit_ids))
            where_conditions.append(
                f"EXISTS (SELECT 1 FROM dbo.APP_IncidentCaseTargetDepartment td_scope"
                f" WHERE td_scope.IncidentRequestCaseID = c.IncidentRequestCaseID"
                f" AND td_scope.DepartmentID IN ({placeholders}))"
            )
            params.extend(list(restrict_to_unit_ids))

    # Domain filter (accepts comma-separated multi-select IDs)
    domain_ids = _parse_multi_int(domain_id)
    if domain_ids:
        where_conditions.append(_in_clause("c.DomainID", domain_ids))
        params.extend(domain_ids)

    # Category filter (accepts comma-separated multi-select IDs)
    category_ids = _parse_multi_int(category_id)
    if category_ids:
        where_conditions.append(_in_clause("c.CategoryID", category_ids))
        params.extend(category_ids)

    # Subcategory filter (accepts comma-separated multi-select IDs)
    subcategory_ids = _parse_multi_int(subcategory_id)
    if subcategory_ids:
        where_conditions.append(_in_clause("c.SubCategoryID", subcategory_ids))
        params.extend(subcategory_ids)

    # Classification filter (accepts comma-separated multi-select IDs)
    classification_ids = _parse_multi_int(classification_id)
    if classification_ids:
        where_conditions.append(_in_clause("c.ClassificationID", classification_ids))
        params.extend(classification_ids)

    # Severity filter (accepts comma-separated multi-select IDs)
    severity_ids = _parse_multi_int(severity_id)
    if severity_ids:
        where_conditions.append(_in_clause("c.SeverityID", severity_ids))
        params.extend(severity_ids)

    # Stage filter
    stage_ids = _parse_multi_int(stage_id)
    if stage_ids:
        where_conditions.append(_in_clause("c.StageID", stage_ids))
        params.extend(stage_ids)

    # Harm level filter (accepts comma-separated multi-select IDs)
    harm_level_ids = _parse_multi_int(harm_level_id)
    if harm_level_ids:
        where_conditions.append(_in_clause("c.HarmLevelID", harm_level_ids))
        params.extend(harm_level_ids)

    # Case status filter (accepts comma-separated multi-select IDs)
    case_status_ids = _parse_multi_int(case_status_id)
    if case_status_ids:
        where_conditions.append(_in_clause("c.CaseStatusID", case_status_ids))
        params.extend(case_status_ids)
    elif tab == 'preparation':
        where_conditions.append("c.CaseStatusID IN (4, 5)")
    elif tab == 'workflow':
        where_conditions.append("c.CaseStatusID NOT IN (4, 5)")

    # Clinical risk type filter (accepts comma-separated multi-select IDs)
    clinical_risk_type_ids = _parse_multi_int(clinical_risk_type_id)
    if clinical_risk_type_ids:
        where_conditions.append(_in_clause("c.ClinicalRiskTypeID", clinical_risk_type_ids))
        params.extend(clinical_risk_type_ids)

    # Feedback intent type filter (accepts comma-separated multi-select IDs)
    feedback_intent_type_ids = _parse_multi_int(feedback_intent_type_id)
    if feedback_intent_type_ids:
        where_conditions.append(_in_clause("c.FeedbackIntentTypeID", feedback_intent_type_ids))
        params.extend(feedback_intent_type_ids)

    # Source filter (accepts comma-separated multi-select IDs)
    source_ids = _parse_multi_int(source_id)
    if source_ids:
        where_conditions.append(_in_clause("c.SourceID", source_ids))
        params.extend(source_ids)

    # Target section filter (exact section ID)
    if target_department_id:
        where_conditions.append("""
            EXISTS (
                SELECT 1 FROM dbo.APP_IncidentCaseTargetDepartment td
                WHERE td.IncidentRequestCaseID = c.IncidentRequestCaseID
                  AND td.DepartmentID = ?
            )
        """)
        params.append(target_department_id)

    # Target department filter (sections belonging to this department, Type=2)
    if target_dept_parent_id:
        where_conditions.append("""
            EXISTS (
                SELECT 1 FROM dbo.APP_IncidentCaseTargetDepartment td
                INNER JOIN AdminsrationUnit s ON s.UniqueID = td.DepartmentID
                WHERE td.IncidentRequestCaseID = c.IncidentRequestCaseID
                  AND s.ParentID = ?
            )
        """)
        params.append(target_dept_parent_id)

    # Target administration filter — match sections directly under this admin OR via an intermediate division
    if target_admin_id:
        where_conditions.append("""
            EXISTS (
                SELECT 1 FROM dbo.APP_IncidentCaseTargetDepartment td
                INNER JOIN AdminsrationUnit s ON s.UniqueID = td.DepartmentID
                LEFT JOIN AdminsrationUnit p ON p.UniqueID = s.ParentID
                WHERE td.IncidentRequestCaseID = c.IncidentRequestCaseID
                  AND (s.ParentID = ? OR p.ParentID = ?)
            )
        """)
        params.extend([target_admin_id, target_admin_id])
    
    # Date filters
    if year:
        where_conditions.append("YEAR(c.FeedbackRecievedDate) = ?")
        params.append(year)
    
    if month:
        where_conditions.append("MONTH(c.FeedbackRecievedDate) = ?")
        params.append(month)
    
    if start_date:
        where_conditions.append("c.FeedbackRecievedDate >= ?")
        params.append(start_date)
    
    if end_date:
        where_conditions.append("c.FeedbackRecievedDate <= ?")
        params.append(end_date)
    
    # Combine WHERE conditions
    where_clause = ""
    if where_conditions:
        where_clause = "WHERE " + " AND ".join(where_conditions)
    
    # Calculate offset for pagination
    offset = (page - 1) * page_size
    
    # Build ORDER BY clause from the trusted whitelist. A stable tiebreaker (case ID)
    # is appended whenever sorting by a non-unique field, so OFFSET/FETCH pagination
    # returns deterministic, non-overlapping pages across requests.
    sort_field_expr = SORT_FIELD_MAP[sort_by]
    sort_dir = sort_order.upper()
    id_expr = SORT_FIELD_MAP["IncidentRequestCaseID"]
    if sort_field_expr == id_expr:
        order_by_clause = f"ORDER BY {sort_field_expr} {sort_dir}"
    else:
        order_by_clause = f"ORDER BY {sort_field_expr} {sort_dir}, {id_expr} ASC"

    # SQL query for fetching records
    query = f"""
        SELECT
            c.IncidentRequestCaseID as id,
            c.IncidentRequestCaseID as complaint_number,
            c.incident_id,
            i.incident_number,
            LEFT(c.ComplaintText, 150) as complaint_summary,
            c.ComplaintText as complaint_text,
            c.FeedbackRecievedDate as received_date,
            c.IncidentDate as incident_date,
            c.CreatedAt as created_at,
            subcase_pub.CreatedAt as publication_date,
            c.PatientName as patient_name,

            -- Issuing organizational unit
            c.IssuingOrgUnitID as issuing_org_unit_id,
            org_unit.Name as issuing_org_unit_name,

            -- Target department(s) — aggregated, for Operational Views
            target_depts.target_dept_names as target_department_name,

            -- Domain
            c.DomainID as domain_id,
            domain.DomainName as domain_name,
            
            -- Category
            c.CategoryID as category_id,
            category.CategoryName as category_name,
            
            -- SubCategory
            c.SubCategoryID as subcategory_id,
            subcategory.SubCategoryName as subcategory_name,
            
            -- Classification
            c.ClassificationID as classification_id,
            classification.Classification_AR as classification_name,
            
            -- Severity
            c.SeverityID as severity_id,
            severity.SeverityName as severity_name,
            
            -- Stage
            c.StageID as stage_id,
            stage.StageName as stage_name,
            
            -- Harm level
            c.HarmLevelID as harm_level_id,
            harm.HarmLevel as harm_level,
            
            -- Case Status
            c.CaseStatusID as case_status_id,
            status.Name as status_name,
            
            -- Building
            c.BuildingID as building_id,
            building.BuildingName as building_name,
            
            -- Risk and Intent Types
            c.ClinicalRiskTypeID as clinical_risk_type_id,
            clinical_risk.Name as clinical_risk_type_name,
            c.FeedbackIntentTypeID as feedback_intent_type_id,
            feedback_intent.NameEn as feedback_intent_type_name,
            
            -- Source
            c.SourceID as source_id,
            source.SourceName as source_name,
            
            -- Explanation Status
            c.ExplanationStatusID as explanation_status_id,
            explanation_status.StatusName as explanation_status_name,
            
            -- Record type (1=Complaint, 2=Notice)
            c.RecordTypeID as record_type_id,

            -- Subcase answers (most recent subcase per case)
            subcase_ans.SectionExplanationText as section_answer,
            subcase_ans.DepartmentExplanationText as department_answer,
            subcase_ans.AdministrationExplanationText as administration_answer,

            -- Customer Service Decision (most recent subcase's Patient Services decision)
            subcase_ans.PatientServicesDecisionText as customer_service_decision,
            subcase_ans.PatientServicesDecisionAt as customer_service_decision_date,

            -- Per-level entry/deadline timestamps (most recent subcase per case)
            subcase_ans.SectionEntryTimestamp as section_entry,
            subcase_ans.SectionDeadlineAt as section_deadline,
            subcase_ans.DepartmentEntryTimestamp as department_entry,
            subcase_ans.DepartmentDeadlineAt as department_deadline,
            subcase_ans.AdministrationEntryTimestamp as administration_entry,
            subcase_ans.AdministrationDeadlineAt as administration_deadline,

            -- RCA Replies: selected Cause -> Corrective Action pairs across the case's subcase(s)
            rca_replies.rca_replies_text as rca_replies,

            -- Other fields
            c.ImmediateAction as immediate_action,
            c.TakenAction as taken_action,
            c.isINPatient as is_inpatient,
            c.CreatedByUserID as created_by_user_id,

            -- Morbidity flag (for Universal Case Theme clinical indicator)
            c.IsMorbidity as is_morbidity,

            -- Recently Edited
            c.UpdatedAt as last_edited

        FROM dbo.APP_IncidentCase c
        LEFT JOIN AdminsrationUnit org_unit ON c.IssuingOrgUnitID = org_unit.UniqueID
        LEFT JOIN APP_LOOKUP_DOMAIN domain ON c.DomainID = domain.DomainID
        LEFT JOIN APP_LOOKUP_CATEGORY category ON c.CategoryID = category.CategoryID
        LEFT JOIN APP_LOOKUP_SUBCATEGORY subcategory ON c.SubCategoryID = subcategory.SubCategoryID
        LEFT JOIN APP_LOOKUP_CLASSIFICATION classification ON c.ClassificationID = classification.ClassificationID
        LEFT JOIN APP_LOOKUP_SEVERITY severity ON c.SeverityID = severity.SeverityID
        LEFT JOIN APP_LOOKUP_CASE_STAGE stage ON c.StageID = stage.StageID
        LEFT JOIN APP_LOOKUP_HARM_LEVEL harm ON c.HarmLevelID = harm.HarmID
        LEFT JOIN APP_LOOKUP_CASE_STATUS status ON c.CaseStatusID = status.CaseStatusID
        LEFT JOIN APP_LOOKUP_BUILDING building ON c.BuildingID = building.BuildingID
        LEFT JOIN APP_LOOKUP_CLINICAL_RISK_TYPE clinical_risk ON c.ClinicalRiskTypeID = clinical_risk.ClinicalRiskTypeID
        LEFT JOIN APP_LOOKUP_FEEDBACK_INTENT_TYPE feedback_intent ON c.FeedbackIntentTypeID = feedback_intent.FeedbackIntentTypeID
        LEFT JOIN APP_LOOKUP_SOURCE source ON c.SourceID = source.SourceID
        LEFT JOIN APP_LOOKUP_EXPLANATION_STATUS explanation_status ON c.ExplanationStatusID = explanation_status.StatusID
        LEFT JOIN dbo.APP_Incident i ON c.incident_id = i.incident_id
        OUTER APPLY (
            SELECT TOP 1
                SectionExplanationText,
                DepartmentExplanationText,
                AdministrationExplanationText,
                PatientServicesDecisionText,
                PatientServicesDecisionAt,
                SectionEntryTimestamp,
                SectionDeadlineAt,
                DepartmentEntryTimestamp,
                DepartmentDeadlineAt,
                AdministrationEntryTimestamp,
                AdministrationDeadlineAt
            FROM dbo.APP_AdministrativeSubcase
            WHERE IncidentRequestCaseID = c.IncidentRequestCaseID
            ORDER BY CreatedAt DESC
        ) subcase_ans
        OUTER APPLY (
            -- Aggregate selected RCA Cause -> Corrective Action pairs across all
            -- of this case's subcases (a case may target more than one subcase).
            SELECT STRING_AGG(
                CONCAT(
                    N'Cause: ', ISNULL(cause.SuggestionTextAr, cause.SuggestionTextEn),
                    N' -> Action: ', ISNULL(action.SuggestionTextAr, action.SuggestionTextEn)
                ),
                N' | '
            ) as rca_replies_text
            FROM dbo.APP_AdministrativeSubcase sub
            JOIN dbo.APP_SubcaseRCASuggestionSelection sel ON sel.SubcaseID = sub.SubcaseID
            JOIN dbo.APP_RCASuggestion cause ON cause.SuggestionID = sel.SuggestionID AND cause.SuggestionType = 'CAUSE'
            JOIN dbo.APP_RCASuggestion action ON action.SuggestionID = cause.PairedSuggestionID
            WHERE sub.IncidentRequestCaseID = c.IncidentRequestCaseID
        ) rca_replies
        OUTER APPLY (
            SELECT STRING_AGG(au.Name, N', ') as target_dept_names
            FROM dbo.APP_IncidentCaseTargetDepartment td
            JOIN dbo.AdminsrationUnit au ON td.DepartmentID = au.UniqueID
            WHERE td.IncidentRequestCaseID = c.IncidentRequestCaseID
        ) target_depts
        OUTER APPLY (
            -- Publication Date: the moment this case first entered workflow
            -- (earliest subcase CreatedAt). NULL for cases never published.
            SELECT TOP 1 CreatedAt
            FROM dbo.APP_AdministrativeSubcase
            WHERE IncidentRequestCaseID = c.IncidentRequestCaseID
            ORDER BY CreatedAt ASC
        ) subcase_pub
        {where_clause}
        {order_by_clause}
        OFFSET ? ROWS
        FETCH NEXT ? ROWS ONLY
    """
    
    # Add pagination params
    params.extend([offset, page_size])
    
    # Get total count
    count_query = f"""
        SELECT COUNT(*) as total
        FROM dbo.APP_IncidentCase c
        LEFT JOIN dbo.APP_Incident i ON c.incident_id = i.incident_id
        {where_clause}
    """
    
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        # Execute count query
        cursor.execute(count_query, params[:-2])  # Exclude offset and page_size
        total_records = cursor.fetchone().total
        
        # Execute main query
        cursor.execute(query, params)
        columns = [column[0] for column in cursor.description]
        
        raw_rows = cursor.fetchall()
        complaints = []
        for row in raw_rows:
            complaint = dict(zip(columns, row))

            if complaint.get('received_date'):
                complaint['received_date'] = complaint['received_date'].strftime('%Y-%m-%d')
            if complaint.get('incident_date'):
                complaint['incident_date'] = complaint['incident_date'].strftime('%Y-%m-%d')
            if complaint.get('created_at'):
                complaint['created_at'] = complaint['created_at'].isoformat()
            if complaint.get('publication_date'):
                complaint['publication_date'] = complaint['publication_date'].isoformat()
            if complaint.get('last_edited'):
                complaint['last_edited'] = complaint['last_edited'].isoformat()
            if complaint.get('customer_service_decision_date'):
                complaint['customer_service_decision_date'] = complaint['customer_service_decision_date'].isoformat()
            for level_field in ('section_entry', 'section_deadline', 'department_entry', 'department_deadline', 'administration_entry', 'administration_deadline'):
                if complaint.get(level_field):
                    complaint[level_field] = complaint[level_field].isoformat()

            if complaint.get('received_date'):
                try:
                    received = datetime.strptime(complaint['received_date'], '%Y-%m-%d')
                    complaint['days_open'] = (datetime.now() - received).days
                except:
                    complaint['days_open'] = None
            else:
                complaint['days_open'] = None

            complaints.append(complaint)

        # Batch fetch all workflow statuses in one query instead of N queries
        case_ids = [c['id'] for c in complaints if c.get('id')]
        workflow_map = _get_workflow_statuses_batch(case_ids, cursor)
        satisfaction_map = get_satisfactions_by_cases(case_ids)
        for complaint in complaints:
            wf = workflow_map.get(complaint.get('id'))
            complaint['workflow_status'] = wf
            complaint['case_status_name'] = complaint.get('status_name')
            complaint['is_late'] = bool(wf and wf.get('is_late', False))

            satisfaction = satisfaction_map.get(complaint.get('id'))
            complaint['satisfaction_status_name'] = satisfaction.get('satisfaction_status_name_en') if satisfaction else None
            complaint['satisfaction_date'] = satisfaction.get('feedback_datetime') if satisfaction else None
        
        total_pages = (total_records + page_size - 1) // page_size
        
        # Build filters_applied dict
        filters_applied = {
            'search': search,
            'issuing_org_unit_id': issuing_org_unit_id,
            'domain_id': domain_id,
            'category_id': category_id,
            'subcategory_id': subcategory_id,
            'classification_id': classification_id,
            'severity_id': severity_id,
            'stage_id': stage_id,
            'harm_level_id': harm_level_id,
            'case_status_id': case_status_id,
            'clinical_risk_type_id': clinical_risk_type_id,
            'feedback_intent_type_id': feedback_intent_type_id,
            'source_id': source_id,
            'year': year,
            'month': month,
            'start_date': start_date,
            'end_date': end_date
        }

        return {
            'complaints': complaints,
            'pagination': {
                'page': page,
                'page_size': page_size,
                'total_records': total_records,
                'total_pages': total_pages
            },
            'filters_applied': filters_applied,
            'view': view
        }
        
    finally:
        cursor.close()
        conn.close()


def get_filter_options(include_counts: bool = False) -> Dict[str, List[Dict[str, Any]]]:
    """
    Fetch available filter options for dropdown population.
    
    Args:
        include_counts: If True, include record count for each option.
        
    Returns:
        Dictionary with arrays for organizational units, domains, categories, severities, etc.
    """
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        result = {}
        
        # Issuing Organizational Units
        org_query = """
            SELECT DISTINCT
                u.UniqueID as id,
                u.Name as name,
                u.ParentID as parent_id
                """ + (", COUNT(c.IncidentRequestCaseID) as count" if include_counts else "") + """
            FROM AdminsrationUnit u
            """ + ("LEFT JOIN APP_IncidentCase c ON u.UniqueID = c.IssuingOrgUnitID" if include_counts else "") + """
            WHERE u.Frozen = 0
            """ + ("GROUP BY u.UniqueID, u.Name, u.ParentID" if include_counts else "") + """
            ORDER BY u.Name
        """
        cursor.execute(org_query)
        result['issuing_org_units'] = [dict(zip([col[0] for col in cursor.description], row)) for row in cursor.fetchall()]
        
        # Domains
        domain_query = """
            SELECT 
                d.DomainID as id,
                d.DomainName as name
                """ + (", COUNT(c.IncidentRequestCaseID) as count" if include_counts else "") + """
            FROM APP_LOOKUP_DOMAIN d
            """ + ("LEFT JOIN APP_IncidentCase c ON d.DomainID = c.DomainID" if include_counts else "") + """
            """ + ("GROUP BY d.DomainID, d.DomainName" if include_counts else "") + """
            ORDER BY d.DomainName
        """
        cursor.execute(domain_query)
        result['domains'] = [dict(zip([col[0] for col in cursor.description], row)) for row in cursor.fetchall()]
        
        # Categories
        category_query = """
            SELECT 
                c.CategoryID as id,
                c.CategoryName as name,
                c.DomainID as domain_id
                """ + (", COUNT(ic.IncidentRequestCaseID) as count" if include_counts else "") + """
            FROM APP_LOOKUP_CATEGORY c
            """ + ("LEFT JOIN APP_IncidentCase ic ON c.CategoryID = ic.CategoryID" if include_counts else "") + """
            """ + ("GROUP BY c.CategoryID, c.CategoryName, c.DomainID" if include_counts else "") + """
            ORDER BY c.CategoryName
        """
        cursor.execute(category_query)
        result['categories'] = [dict(zip([col[0] for col in cursor.description], row)) for row in cursor.fetchall()]
        
        # Severities
        severity_query = """
            SELECT 
                s.SeverityID as id,
                s.SeverityName as name
                """ + (", COUNT(c.IncidentRequestCaseID) as count" if include_counts else "") + """
            FROM APP_LOOKUP_SEVERITY s
            """ + ("LEFT JOIN APP_IncidentCase c ON s.SeverityID = c.SeverityID" if include_counts else "") + """
            WHERE s.IsActive = 1
            """ + ("GROUP BY s.SeverityID, s.SeverityName" if include_counts else "") + """
            ORDER BY s.SeverityOrder
        """
        cursor.execute(severity_query)
        result['severities'] = [dict(zip([col[0] for col in cursor.description], row)) for row in cursor.fetchall()]
        
        # Stages
        stage_query = """
            SELECT 
                s.StageID as id,
                s.StageName as name
                """ + (", COUNT(c.IncidentRequestCaseID) as count" if include_counts else "") + """
            FROM APP_LOOKUP_CASE_STAGE s
            """ + ("LEFT JOIN APP_IncidentCase c ON s.StageID = c.StageID" if include_counts else "") + """
            """ + ("GROUP BY s.StageID, s.StageName" if include_counts else "") + """
            ORDER BY s.StageID
        """
        cursor.execute(stage_query)
        result['stages'] = [dict(zip([col[0] for col in cursor.description], row)) for row in cursor.fetchall()]
        
        # Harm Levels
        harm_query = """
            SELECT 
                h.HarmID as id,
                h.HarmLevel as name
                """ + (", COUNT(c.IncidentRequestCaseID) as count" if include_counts else "") + """
            FROM APP_LOOKUP_HARM_LEVEL h
            """ + ("LEFT JOIN APP_IncidentCase c ON h.HarmID = c.HarmLevelID" if include_counts else "") + """
            """ + ("GROUP BY h.HarmID, h.HarmLevel" if include_counts else "") + """
            ORDER BY h.HarmID
        """
        cursor.execute(harm_query)
        result['harm_levels'] = [dict(zip([col[0] for col in cursor.description], row)) for row in cursor.fetchall()]
        
        # Statuses
        status_query = """
            SELECT 
                s.CaseStatusID as id,
                s.Name as name
                """ + (", COUNT(c.IncidentRequestCaseID) as count" if include_counts else "") + """
            FROM APP_LOOKUP_CASE_STATUS s
            """ + ("LEFT JOIN APP_IncidentCase c ON s.CaseStatusID = c.CaseStatusID" if include_counts else "") + """
            """ + ("GROUP BY s.CaseStatusID, s.Name" if include_counts else "") + """
            WHERE s.IsActive = 1
            ORDER BY s.DisplayOrder
        """
        cursor.execute(status_query)
        result['statuses'] = [dict(zip([col[0] for col in cursor.description], row)) for row in cursor.fetchall()]
        
        # Classifications EN
        classification_en_query = """
            SELECT 
                c.ClassificationID as id,
                c.Classification_AR as name
                """ + (", COUNT(ic.IncidentRequestCaseID) as count" if include_counts else "") + """
            FROM APP_LOOKUP_CLASSIFICATION c
            """ + ("LEFT JOIN APP_IncidentCase ic ON c.ClassificationID = ic.ClassificationID" if include_counts else "") + """
            """ + ("GROUP BY c.ClassificationID, c.Classification_AR" if include_counts else "") + """
            ORDER BY c.Classification_AR
        """
        cursor.execute(classification_en_query)
        result['classifications_en'] = [dict(zip([col[0] for col in cursor.description], row)) for row in cursor.fetchall()]
        
        # Years
        years_query = """
            SELECT DISTINCT YEAR(FeedbackRecievedDate) as year
            FROM APP_IncidentCase
            WHERE FeedbackRecievedDate IS NOT NULL
            ORDER BY year DESC
        """
        cursor.execute(years_query)
        result['years'] = [row[0] for row in cursor.fetchall()]

        # Target Departments (Type=323 administrations + Type=325 divisions — parents of sections)
        cursor.execute("""
            SELECT d.UniqueID as id, d.Name as name, d.ParentID as administration_id, a.Name as administration_name
            FROM AdminsrationUnit d
            LEFT JOIN AdminsrationUnit a ON a.UniqueID = d.ParentID
            WHERE d.Type IN (323, 325) AND d.Frozen = 0
            ORDER BY a.Name, d.Name
        """)
        result['target_departments'] = [dict(zip([c[0] for c in cursor.description], row)) for row in cursor.fetchall()]

        # Target Administrations (Type=323 — top-level)
        cursor.execute("""
            SELECT UniqueID as id, Name as name
            FROM AdminsrationUnit
            WHERE Type = 323 AND Frozen = 0
            ORDER BY Name
        """)
        result['target_administrations'] = [dict(zip([c[0] for c in cursor.description], row)) for row in cursor.fetchall()]

        return result
        
    finally:
        cursor.close()
        conn.close()


def get_complaint_by_id(complaint_id: int) -> Optional[Dict[str, Any]]:
    """
    Fetch full details of a single complaint record.
    
    Args:
        complaint_id: The IncidentRequestCaseID to retrieve.
        
    Returns:
        Dictionary with full complaint details, or None if not found.
    """
    query = """
        SELECT
            c.IncidentRequestCaseID as id,
            c.ComplaintText as complaint_text,
            c.ImmediateAction as immediate_action,
            c.TakenAction as taken_action,
            c.FeedbackRecievedDate as received_date,
            c.IncidentDate as incident_date,
            c.PatientName as patient_name,
            c.CreatedAt as created_at,
            c.UpdatedAt as updated_at,
            c.CreatedByUserID as created_by_user_id,
            c.isINPatient as is_inpatient,
            c.IsMorbidity as is_morbidity,

            -- Incident parent
            c.incident_id as incident_id,
            inc.incident_number as incident_number,

            -- Issuing organizational unit
            c.IssuingOrgUnitID as issuing_org_unit_id,
            org_unit.Name as issuing_org_unit_name,
            
            -- Domain
            c.DomainID as domain_id,
            domain.DomainName as domain_name,
            
            -- Category
            c.CategoryID as category_id,
            category.CategoryName as category_name,
            
            -- SubCategory
            c.SubCategoryID as subcategory_id,
            subcategory.SubCategoryName as subcategory_name,
            
            -- Classification
            c.ClassificationID as classification_id,
            classification.Classification_AR as classification_name,
            
            -- Severity
            c.SeverityID as severity_id,
            severity.SeverityName as severity_name,
            
            -- Stage
            c.StageID as stage_id,
            stage.StageName as stage_name,
            
            -- Harm level
            c.HarmLevelID as harm_level_id,
            harm.HarmLevel as harm_level,
            
            -- Case Status
            c.CaseStatusID as case_status_id,
            status.Name as status_name,
            
            -- Building
            c.BuildingID as building_id,
            building.BuildingName as building_name,
            
            -- Risk and Intent Types
            c.ClinicalRiskTypeID as clinical_risk_type_id,
            clinical_risk.Name as clinical_risk_type_name,
            c.FeedbackIntentTypeID as feedback_intent_type_id,
            feedback_intent.NameEn as feedback_intent_type_name,
            
            -- Source
            c.SourceID as source_id,
            source.SourceName as source_name,
            
            -- Explanation Status
            c.ExplanationStatusID as explanation_status_id,
            explanation_status.StatusName as explanation_status_name,

            -- Record type (1=Complaint, 2=Notice)
            c.RecordTypeID as record_type_id,

            -- Incident parent fields (for draft prefill)
            c.incident_id as incident_id,
            inc.incident_number as incident_number,
            inc.complaint_summary as incident_summary
        FROM dbo.APP_IncidentCase c
        LEFT JOIN AdminsrationUnit org_unit ON c.IssuingOrgUnitID = org_unit.UniqueID
        LEFT JOIN APP_LOOKUP_DOMAIN domain ON c.DomainID = domain.DomainID
        LEFT JOIN APP_LOOKUP_CATEGORY category ON c.CategoryID = category.CategoryID
        LEFT JOIN APP_LOOKUP_SUBCATEGORY subcategory ON c.SubCategoryID = subcategory.SubCategoryID
        LEFT JOIN APP_LOOKUP_CLASSIFICATION classification ON c.ClassificationID = classification.ClassificationID
        LEFT JOIN APP_LOOKUP_SEVERITY severity ON c.SeverityID = severity.SeverityID
        LEFT JOIN APP_LOOKUP_CASE_STAGE stage ON c.StageID = stage.StageID
        LEFT JOIN APP_LOOKUP_HARM_LEVEL harm ON c.HarmLevelID = harm.HarmID
        LEFT JOIN APP_LOOKUP_CASE_STATUS status ON c.CaseStatusID = status.CaseStatusID
        LEFT JOIN APP_LOOKUP_BUILDING building ON c.BuildingID = building.BuildingID
        LEFT JOIN APP_LOOKUP_CLINICAL_RISK_TYPE clinical_risk ON c.ClinicalRiskTypeID = clinical_risk.ClinicalRiskTypeID
        LEFT JOIN APP_LOOKUP_FEEDBACK_INTENT_TYPE feedback_intent ON c.FeedbackIntentTypeID = feedback_intent.FeedbackIntentTypeID
        LEFT JOIN APP_LOOKUP_SOURCE source ON c.SourceID = source.SourceID
        LEFT JOIN APP_LOOKUP_EXPLANATION_STATUS explanation_status ON c.ExplanationStatusID = explanation_status.StatusID
        LEFT JOIN dbo.APP_Incident inc ON c.incident_id = inc.incident_id
        WHERE c.IncidentRequestCaseID = ?
    """
    
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute(query, (complaint_id,))
        row = cursor.fetchone()
        
        if not row:
            return None
        
        columns = [column[0] for column in cursor.description]
        complaint = dict(zip(columns, row))
        
        # Format dates
        if complaint.get('received_date'):
            complaint['received_date'] = complaint['received_date'].strftime('%Y-%m-%d')
        if complaint.get('incident_date'):
            complaint['incident_date'] = complaint['incident_date'].strftime('%Y-%m-%d')
        if complaint.get('created_at'):
            complaint['created_at'] = complaint['created_at'].isoformat()
        if complaint.get('updated_at'):
            complaint['updated_at'] = complaint['updated_at'].isoformat()
        
        # Fetch target departments.
        # Uses OUTER APPLY to walk the hierarchy by Type rather than by position,
        # so both depth-2 (Section directly under Administration) and depth-3
        # (Section → Department → Administration) chains return correct labels.
        # Self-referencing root nodes (ParentID = UniqueID for Administrations)
        # are excluded from the upward walk via the != guard.
        target_dept_query = """
            SELECT
                td.IsPrimary AS is_primary,
                hier.section_id,
                hier.section_name,
                hier.department_id,
                hier.department_name,
                hier.administration_id,
                hier.administration_name
            FROM dbo.APP_IncidentCaseTargetDepartment td
            OUTER APPLY (
                SELECT
                    -- Section (Type=324): only when the stored target is itself a Section
                    CASE WHEN target.Type = 324 THEN target.UniqueID ELSE NULL END AS section_id,
                    CASE WHEN target.Type = 324 THEN target.Name    ELSE NULL END AS section_name,

                    -- Department (Type=325): the stored target if it is a Department,
                    --   or the parent of a Section if that parent is a Department
                    CASE
                        WHEN target.Type = 325 THEN target.UniqueID
                        WHEN target.Type = 324 AND p1.Type = 325 THEN p1.UniqueID
                        ELSE NULL
                    END AS department_id,
                    CASE
                        WHEN target.Type = 325 THEN target.Name
                        WHEN target.Type = 324 AND p1.Type = 325 THEN p1.Name
                        ELSE NULL
                    END AS department_name,

                    -- Administration (Type=323): first ancestor of Type=323 walking upward
                    CASE
                        WHEN target.Type = 323 THEN target.UniqueID
                        WHEN target.Type = 325 AND p1.Type = 323 THEN p1.UniqueID
                        WHEN target.Type = 324 AND p1.Type = 323 THEN p1.UniqueID
                        WHEN target.Type = 324 AND p1.Type = 325 AND p2.Type = 323 THEN p2.UniqueID
                        ELSE NULL
                    END AS administration_id,
                    CASE
                        WHEN target.Type = 323 THEN target.Name
                        WHEN target.Type = 325 AND p1.Type = 323 THEN p1.Name
                        WHEN target.Type = 324 AND p1.Type = 323 THEN p1.Name
                        WHEN target.Type = 324 AND p1.Type = 325 AND p2.Type = 323 THEN p2.Name
                        ELSE NULL
                    END AS administration_name

                FROM dbo.AdminsrationUnit target
                -- p1: direct parent, skip self-referencing roots
                LEFT JOIN dbo.AdminsrationUnit p1
                    ON target.ParentID = p1.UniqueID
                    AND target.ParentID != target.UniqueID
                -- p2: grandparent, skip self-referencing roots
                LEFT JOIN dbo.AdminsrationUnit p2
                    ON p1.ParentID = p2.UniqueID
                    AND p1.ParentID IS NOT NULL
                    AND p1.ParentID != p1.UniqueID
                WHERE target.UniqueID = td.DepartmentID
            ) AS hier
            WHERE td.IncidentRequestCaseID = ?
            ORDER BY td.IsPrimary DESC, td.DepartmentID
        """
        cursor.execute(target_dept_query, (complaint_id,))
        target_dept_rows = cursor.fetchall()
        
        target_departments = []
        for dept_row in target_dept_rows:
            target_departments.append({
                'section_id': dept_row.section_id,
                'section_name': dept_row.section_name,
                'department_id': dept_row.department_id,
                'department_name': dept_row.department_name,
                'administration_id': dept_row.administration_id,
                'administration_name': dept_row.administration_name,
                'is_primary': bool(dept_row.is_primary)
            })
        
        complaint['target_departments'] = target_departments
        
        # Fetch linked employees (supervisors/workers)
        employee_query = f"""
            SELECT 
                e.EmployeeID as employee_id,
                hr.FullName as full_name,
                hr.JobTitle as job_title,
                e.IsPrimary as is_primary
            FROM dbo.APP_IncidentCaseEmployee e
            LEFT JOIN {HR_EMPLOYEES_TABLE} hr ON e.EmployeeID = hr.EmployeeID
            WHERE e.IncidentRequestCaseID = ?
            ORDER BY e.IsPrimary DESC, e.EmployeeID
        """
        cursor.execute(employee_query, (complaint_id,))
        employee_rows = cursor.fetchall()
        
        employees = []
        for emp_row in employee_rows:
            employees.append({
                'employee_id': emp_row.employee_id,
                'full_name': emp_row.full_name,
                'job_title': emp_row.job_title if hasattr(emp_row, 'job_title') else None,
                'is_primary': bool(emp_row.is_primary) if emp_row.is_primary is not None else False
            })
        
        complaint['employees'] = employees
        
        # Fetch linked doctors (check both hospital view and reserve table)
        doctor_query = f"""
            SELECT 
                d.DoctorID as doctor_id,
                COALESCE(doc.Name, reserve.DoctorName) as doctor_name,
                d.IsPrimary as is_primary
            FROM dbo.APP_IncidentCaseDoctor d
            LEFT JOIN {DOCTORS_TABLE} doc ON d.DoctorID = doc.DoctorID
            LEFT JOIN dbo.APP_RESERVE_DOCTOR reserve ON d.DoctorID = reserve.DoctorID
            WHERE d.IncidentRequestCaseID = ?
            ORDER BY d.IsPrimary DESC, d.DoctorID
        """
        cursor.execute(doctor_query, (complaint_id,))
        doctor_rows = cursor.fetchall()
        
        doctors = []
        for doc_row in doctor_rows:
            doctors.append({
                'doctor_id': doc_row.doctor_id,
                'doctor_name': doc_row.doctor_name,
                'is_primary': bool(doc_row.is_primary) if doc_row.is_primary is not None else False
            })
        
        complaint['doctors'] = doctors
        
        return complaint
        
    finally:
        cursor.close()
        conn.close()


def get_complaints_count(
    search: Optional[str] = None,
    issuing_org_unit_id: Optional[int] = None,
    target_department_id: Optional[int] = None,
    target_dept_parent_id: Optional[int] = None,
    target_admin_id: Optional[int] = None,
    domain_id: Optional[str] = None,
    category_id: Optional[str] = None,
    subcategory_id: Optional[str] = None,
    classification_id: Optional[str] = None,
    severity_id: Optional[str] = None,
    stage_id: Optional[str] = None,
    harm_level_id: Optional[str] = None,
    case_status_id: Optional[str] = None,
    clinical_risk_type_id: Optional[str] = None,
    feedback_intent_type_id: Optional[str] = None,
    source_id: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    tab: Optional[str] = None,
    restrict_to_unit_ids: Optional[set] = None,
) -> Dict[str, Any]:
    """
    Get count of complaints matching filters (for export preview).
    
    Returns:
        Dictionary with total_count and filters_applied.
    """
    # Build WHERE clause (same as get_complaints_paginated)
    where_conditions = []
    params = []
    
    if search:
        search = search.strip()
    if search:
        search_condition = """(
            CAST(c.IncidentRequestCaseID AS VARCHAR) LIKE ?
            OR c.PatientName LIKE ?
            OR c.ComplaintText LIKE ?
            OR i.incident_number LIKE ?
        )"""
        where_conditions.append(search_condition)
        search_param = f"%{search}%"
        params.extend([search_param, search_param, search_param, search_param])

    if issuing_org_unit_id:
        where_conditions.append("c.IssuingOrgUnitID = ?")
        params.append(issuing_org_unit_id)

    if restrict_to_unit_ids is not None:
        if not restrict_to_unit_ids:
            where_conditions.append("1=0")
        else:
            placeholders = ",".join("?" * len(restrict_to_unit_ids))
            where_conditions.append(
                f"EXISTS (SELECT 1 FROM dbo.APP_IncidentCaseTargetDepartment td_scope"
                f" WHERE td_scope.IncidentRequestCaseID = c.IncidentRequestCaseID"
                f" AND td_scope.DepartmentID IN ({placeholders}))"
            )
            params.extend(list(restrict_to_unit_ids))

    if target_department_id:
        where_conditions.append("""
            EXISTS (SELECT 1 FROM dbo.APP_IncidentCaseTargetDepartment td
                    WHERE td.IncidentRequestCaseID = c.IncidentRequestCaseID AND td.DepartmentID = ?)
        """)
        params.append(target_department_id)

    if target_dept_parent_id:
        where_conditions.append("""
            EXISTS (SELECT 1 FROM dbo.APP_IncidentCaseTargetDepartment td
                    INNER JOIN AdminsrationUnit s ON s.UniqueID = td.DepartmentID
                    WHERE td.IncidentRequestCaseID = c.IncidentRequestCaseID AND s.ParentID = ?)
        """)
        params.append(target_dept_parent_id)

    if target_admin_id:
        where_conditions.append("""
            EXISTS (SELECT 1 FROM dbo.APP_IncidentCaseTargetDepartment td
                    INNER JOIN AdminsrationUnit s ON s.UniqueID = td.DepartmentID
                    INNER JOIN AdminsrationUnit d ON d.UniqueID = s.ParentID
                    WHERE td.IncidentRequestCaseID = c.IncidentRequestCaseID AND d.ParentID = ?)
        """)
        params.append(target_admin_id)

    domain_ids = _parse_multi_int(domain_id)
    if domain_ids:
        where_conditions.append(_in_clause("c.DomainID", domain_ids))
        params.extend(domain_ids)

    category_ids = _parse_multi_int(category_id)
    if category_ids:
        where_conditions.append(_in_clause("c.CategoryID", category_ids))
        params.extend(category_ids)

    subcategory_ids = _parse_multi_int(subcategory_id)
    if subcategory_ids:
        where_conditions.append(_in_clause("c.SubCategoryID", subcategory_ids))
        params.extend(subcategory_ids)

    classification_ids = _parse_multi_int(classification_id)
    if classification_ids:
        where_conditions.append(_in_clause("c.ClassificationID", classification_ids))
        params.extend(classification_ids)

    severity_ids = _parse_multi_int(severity_id)
    if severity_ids:
        where_conditions.append(_in_clause("c.SeverityID", severity_ids))
        params.extend(severity_ids)

    stage_ids = _parse_multi_int(stage_id)
    if stage_ids:
        where_conditions.append(_in_clause("c.StageID", stage_ids))
        params.extend(stage_ids)

    harm_level_ids = _parse_multi_int(harm_level_id)
    if harm_level_ids:
        where_conditions.append(_in_clause("c.HarmLevelID", harm_level_ids))
        params.extend(harm_level_ids)

    case_status_ids = _parse_multi_int(case_status_id)
    if case_status_ids:
        where_conditions.append(_in_clause("c.CaseStatusID", case_status_ids))
        params.extend(case_status_ids)
    elif tab == 'preparation':
        where_conditions.append("c.CaseStatusID IN (4, 5)")
    elif tab == 'workflow':
        where_conditions.append("c.CaseStatusID NOT IN (4, 5)")

    clinical_risk_type_ids = _parse_multi_int(clinical_risk_type_id)
    if clinical_risk_type_ids:
        where_conditions.append(_in_clause("c.ClinicalRiskTypeID", clinical_risk_type_ids))
        params.extend(clinical_risk_type_ids)

    feedback_intent_type_ids = _parse_multi_int(feedback_intent_type_id)
    if feedback_intent_type_ids:
        where_conditions.append(_in_clause("c.FeedbackIntentTypeID", feedback_intent_type_ids))
        params.extend(feedback_intent_type_ids)

    source_ids = _parse_multi_int(source_id)
    if source_ids:
        where_conditions.append(_in_clause("c.SourceID", source_ids))
        params.extend(source_ids)

    if year:
        where_conditions.append("YEAR(c.FeedbackRecievedDate) = ?")
        params.append(year)

    if month:
        where_conditions.append("MONTH(c.FeedbackRecievedDate) = ?")
        params.append(month)

    if start_date:
        where_conditions.append("c.FeedbackRecievedDate >= ?")
        params.append(start_date)

    if end_date:
        where_conditions.append("c.FeedbackRecievedDate <= ?")
        params.append(end_date)

    where_clause = ""
    if where_conditions:
        where_clause = "WHERE " + " AND ".join(where_conditions)

    query = f"""
        SELECT COUNT(*) as total
        FROM dbo.APP_IncidentCase c
        LEFT JOIN dbo.APP_Incident i ON c.incident_id = i.incident_id
        {where_clause}
    """

    conn = get_connection()
    cursor = conn.cursor()

    try:
        cursor.execute(query, params)
        total_count = cursor.fetchone().total

        filters_applied = {
            'search': search,
            'issuing_org_unit_id': issuing_org_unit_id,
            'domain_id': domain_id,
            'category_id': category_id,
            'subcategory_id': subcategory_id,
            'classification_id': classification_id,
            'severity_id': severity_id,
            'stage_id': stage_id,
            'harm_level_id': harm_level_id,
            'case_status_id': case_status_id,
            'year': year,
            'month': month,
            'start_date': start_date,
            'end_date': end_date
        }
        
        return {
            'total_count': total_count,
            'filters_applied': filters_applied
        }
        
    finally:
        cursor.close()
        conn.close()


def export_complaints_excel(
    search: Optional[str] = None,
    issuing_org_unit_id: Optional[int] = None,
    target_department_id: Optional[int] = None,
    target_dept_parent_id: Optional[int] = None,
    target_admin_id: Optional[int] = None,
    domain_id: Optional[str] = None,
    category_id: Optional[str] = None,
    subcategory_id: Optional[str] = None,
    classification_id: Optional[str] = None,
    severity_id: Optional[str] = None,
    stage_id: Optional[str] = None,
    harm_level_id: Optional[str] = None,
    case_status_id: Optional[str] = None,
    clinical_risk_type_id: Optional[str] = None,
    feedback_intent_type_id: Optional[str] = None,
    source_id: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    sort_by: str = "FeedbackRecievedDate",
    sort_order: str = "desc",
    tab: Optional[str] = None,
    restrict_to_unit_ids: Optional[set] = None,
) -> BytesIO:
    """
    Export filtered complaints as Excel file.
    
    Returns:
        BytesIO object containing the Excel file.
    """
    # Build WHERE clause (same logic as get_complaints_paginated)
    where_conditions = []
    params = []
    
    if search:
        search_condition = """(
            CAST(c.IncidentRequestCaseID AS VARCHAR) LIKE ? 
            OR c.PatientName LIKE ? 
            OR c.ComplaintText LIKE ?
        )"""
        where_conditions.append(search_condition)
        search_param = f"%{search}%"
        params.extend([search_param, search_param, search_param])
    
    if issuing_org_unit_id:
        where_conditions.append("c.IssuingOrgUnitID = ?")
        params.append(issuing_org_unit_id)

    if restrict_to_unit_ids is not None:
        if not restrict_to_unit_ids:
            where_conditions.append("1=0")
        else:
            placeholders = ",".join("?" * len(restrict_to_unit_ids))
            where_conditions.append(
                f"EXISTS (SELECT 1 FROM dbo.APP_IncidentCaseTargetDepartment td_scope"
                f" WHERE td_scope.IncidentRequestCaseID = c.IncidentRequestCaseID"
                f" AND td_scope.DepartmentID IN ({placeholders}))"
            )
            params.extend(list(restrict_to_unit_ids))

    if target_department_id:
        where_conditions.append("""
            EXISTS (
                SELECT 1 FROM dbo.APP_IncidentCaseTargetDepartment td
                WHERE td.IncidentRequestCaseID = c.IncidentRequestCaseID
                  AND td.DepartmentID = ?
            )
        """)
        params.append(target_department_id)

    if target_dept_parent_id:
        where_conditions.append("""
            EXISTS (
                SELECT 1 FROM dbo.APP_IncidentCaseTargetDepartment td
                INNER JOIN AdminsrationUnit s ON s.UniqueID = td.DepartmentID
                WHERE td.IncidentRequestCaseID = c.IncidentRequestCaseID
                  AND s.ParentID = ?
            )
        """)
        params.append(target_dept_parent_id)

    if target_admin_id:
        where_conditions.append("""
            EXISTS (
                SELECT 1 FROM dbo.APP_IncidentCaseTargetDepartment td
                INNER JOIN AdminsrationUnit s ON s.UniqueID = td.DepartmentID
                LEFT JOIN AdminsrationUnit p ON p.UniqueID = s.ParentID
                WHERE td.IncidentRequestCaseID = c.IncidentRequestCaseID
                  AND (s.ParentID = ? OR p.ParentID = ?)
            )
        """)
        params.extend([target_admin_id, target_admin_id])

    domain_ids = _parse_multi_int(domain_id)
    if domain_ids:
        where_conditions.append(_in_clause("c.DomainID", domain_ids))
        params.extend(domain_ids)

    category_ids = _parse_multi_int(category_id)
    if category_ids:
        where_conditions.append(_in_clause("c.CategoryID", category_ids))
        params.extend(category_ids)

    subcategory_ids = _parse_multi_int(subcategory_id)
    if subcategory_ids:
        where_conditions.append(_in_clause("c.SubCategoryID", subcategory_ids))
        params.extend(subcategory_ids)

    classification_ids = _parse_multi_int(classification_id)
    if classification_ids:
        where_conditions.append(_in_clause("c.ClassificationID", classification_ids))
        params.extend(classification_ids)

    severity_ids = _parse_multi_int(severity_id)
    if severity_ids:
        where_conditions.append(_in_clause("c.SeverityID", severity_ids))
        params.extend(severity_ids)

    stage_ids = _parse_multi_int(stage_id)
    if stage_ids:
        where_conditions.append(_in_clause("c.StageID", stage_ids))
        params.extend(stage_ids)

    harm_level_ids = _parse_multi_int(harm_level_id)
    if harm_level_ids:
        where_conditions.append(_in_clause("c.HarmLevelID", harm_level_ids))
        params.extend(harm_level_ids)

    case_status_ids = _parse_multi_int(case_status_id)
    if case_status_ids:
        where_conditions.append(_in_clause("c.CaseStatusID", case_status_ids))
        params.extend(case_status_ids)
    elif tab == 'preparation':
        where_conditions.append("c.CaseStatusID IN (4, 5)")
    elif tab == 'workflow':
        where_conditions.append("c.CaseStatusID NOT IN (4, 5)")

    clinical_risk_type_ids = _parse_multi_int(clinical_risk_type_id)
    if clinical_risk_type_ids:
        where_conditions.append(_in_clause("c.ClinicalRiskTypeID", clinical_risk_type_ids))
        params.extend(clinical_risk_type_ids)

    feedback_intent_type_ids = _parse_multi_int(feedback_intent_type_id)
    if feedback_intent_type_ids:
        where_conditions.append(_in_clause("c.FeedbackIntentTypeID", feedback_intent_type_ids))
        params.extend(feedback_intent_type_ids)

    source_ids = _parse_multi_int(source_id)
    if source_ids:
        where_conditions.append(_in_clause("c.SourceID", source_ids))
        params.extend(source_ids)

    if year:
        where_conditions.append("YEAR(c.FeedbackRecievedDate) = ?")
        params.append(year)

    if month:
        where_conditions.append("MONTH(c.FeedbackRecievedDate) = ?")
        params.append(month)

    if start_date:
        where_conditions.append("c.FeedbackRecievedDate >= ?")
        params.append(start_date)

    if end_date:
        where_conditions.append("c.FeedbackRecievedDate <= ?")
        params.append(end_date)

    where_clause = ""
    if where_conditions:
        where_clause = "WHERE " + " AND ".join(where_conditions)

    # Build ORDER BY — export respects active sort.
    # Same sort keys as get_complaints_paginated's SORT_FIELD_MAP, but resolved against
    # this query's own table aliases (export uses a separate FROM/JOIN structure).
    # FeedbackIntentTypeName has no name join here (export uses an inline CASE WHEN for
    # display instead) so it falls back to the raw FK — numeric, not alphabetical, but
    # consistent ordering is still preserved.
    EXPORT_SORT_FIELD_MAP = {
        "FeedbackRecievedDate": "c.FeedbackRecievedDate",
        "IncidentRequestCaseID": "CAST(c.IncidentRequestCaseID AS INT)",
        "id": "CAST(c.IncidentRequestCaseID AS INT)",
        "CreatedAt": "c.CreatedAt",
        "UpdatedAt": "c.UpdatedAt",
        "PatientName": "c.PatientName",
        "SeverityID": "severity.SeverityOrder",
        "HarmSeverityOrder": "harm.SeverityOrder",
        "IncidentNumber": "inc.incident_number",
        "SourceName": "source.SourceNameAr",
        "FeedbackIntentTypeName": "c.FeedbackIntentTypeID",
        "DomainName": "domain.DomainName",
        "CategoryName": "category.CategoryName",
        "SubCategoryName": "subcat.SubCategoryName",
        "ClassificationName": "classif.Classification_AR",
        "StatusDisplayOrder": "status.DisplayOrder",
        "TargetDepartmentName": "target_depts.target_dept_names",
    }
    sort_field_expr = EXPORT_SORT_FIELD_MAP.get(sort_by, "c.FeedbackRecievedDate")
    sort_dir = "DESC" if sort_order.lower() != "asc" else "ASC"
    id_expr = EXPORT_SORT_FIELD_MAP["IncidentRequestCaseID"]
    if sort_field_expr == id_expr:
        order_by_clause = f"ORDER BY {sort_field_expr} {sort_dir}"
    else:
        order_by_clause = f"ORDER BY {sort_field_expr} {sort_dir}, {id_expr} ASC"

    # Query to fetch all matching records with all required fields
    # Note: Using OUTER APPLY with STRING_AGG for target departments (multiple per case)
    query = f"""
        SELECT
            c.FeedbackRecievedDate as received_date,
            c.IncidentRequestCaseID as complaint_number,
            inc.incident_number as incident_number,
            c.PatientName as patient_name,
            issuing_org.Name as issuing_org_unit_name,
            target_depts.target_dept_names as concerned_org_unit_name,
            source.SourceNameAr as source_name,
            CASE
                WHEN c.FeedbackIntentTypeID = 1 THEN N'شكوى'
                WHEN c.FeedbackIntentTypeID = 2 THEN N'ملاحظة'
                WHEN c.FeedbackIntentTypeID = 3 THEN N'اقتراح'
                WHEN c.FeedbackIntentTypeID = 4 THEN N'استفسار'
                ELSE N'غير محدد'
            END as feedback_type,
            domain.DomainName as domain_name,
            category.CategoryName as category_name,
            subcat.SubCategoryName as subcategory_name,
            classif.Classification_AR as classification_name,
            c.ComplaintText as complaint_text,
            c.ImmediateAction as immediate_action,
            c.TakenAction as taken_action,
            severity.SeverityName as severity_name,
            stage.StageName as stage_name,
            harm.HarmLevel as harm_level,
            status.Name as status_name,
            risk_type.Name as feedback_risk_type,
            export_subcase.SectionExplanationText as section_answer,
            export_subcase.DepartmentExplanationText as department_answer,
            export_subcase.AdministrationExplanationText as administration_answer,
            c.UpdatedAt as last_edited
        FROM dbo.APP_IncidentCase c
        LEFT JOIN AdminsrationUnit issuing_org ON c.IssuingOrgUnitID = issuing_org.UniqueID
        OUTER APPLY (
            SELECT STRING_AGG(au.Name, N'، ') as target_dept_names
            FROM dbo.APP_IncidentCaseTargetDepartment td
            JOIN dbo.AdminsrationUnit au ON td.DepartmentID = au.UniqueID
            WHERE td.IncidentRequestCaseID = c.IncidentRequestCaseID
        ) target_depts
        OUTER APPLY (
            SELECT TOP 1
                SectionExplanationText,
                DepartmentExplanationText,
                AdministrationExplanationText
            FROM dbo.APP_AdministrativeSubcase
            WHERE IncidentRequestCaseID = c.IncidentRequestCaseID
            ORDER BY CreatedAt DESC
        ) export_subcase
        LEFT JOIN APP_LOOKUP_DOMAIN domain ON c.DomainID = domain.DomainID
        LEFT JOIN APP_LOOKUP_CATEGORY category ON c.CategoryID = category.CategoryID
        LEFT JOIN APP_LOOKUP_SUBCATEGORY subcat ON c.SubCategoryID = subcat.SubCategoryID
        LEFT JOIN APP_LOOKUP_CLASSIFICATION classif ON c.ClassificationID = classif.ClassificationID
        LEFT JOIN APP_LOOKUP_SOURCE source ON c.SourceID = source.SourceID
        LEFT JOIN APP_LOOKUP_SEVERITY severity ON c.SeverityID = severity.SeverityID
        LEFT JOIN APP_LOOKUP_CASE_STAGE stage ON c.StageID = stage.StageID
        LEFT JOIN APP_LOOKUP_HARM_LEVEL harm ON c.HarmLevelID = harm.HarmID
        LEFT JOIN APP_LOOKUP_CASE_STATUS status ON c.CaseStatusID = status.CaseStatusID
        LEFT JOIN APP_LOOKUP_CLINICAL_RISK_TYPE risk_type ON c.ClinicalRiskTypeID = risk_type.ClinicalRiskTypeID
        LEFT JOIN dbo.APP_Incident inc ON c.incident_id = inc.incident_id
        {where_clause}
        {order_by_clause}
    """
    
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute(query, params)
        rows = cursor.fetchall()
        columns = [column[0] for column in cursor.description]
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Complaints Export"
        
        # Set worksheet to Right-to-Left for Arabic text
        ws.sheet_view.rightToLeft = True
        
        # Define header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write headers (Arabic translations) - Exact order as specified
        headers_arabic = {
            'received_date': 'تاريخ تلقي الملاحظة',
            'complaint_number': 'رقم الحالة',
            'incident_number': 'رقم الحادثة',
            'patient_name': 'اسم المريض',
            'issuing_org_unit_name': 'قسم الصادر',
            'concerned_org_unit_name': 'قسم المعني',
            'source_name': 'المصدر',
            'feedback_type': 'النوع (Feedback Type)',
            'domain_name': 'Domain',
            'category_name': 'Category',
            'subcategory_name': 'SubCategory',
            'classification_name': 'New Classification',
            'complaint_text': 'محتوى الشكوى (Raw Content)',
            'immediate_action': 'Immediate Action',
            'taken_action': 'الإجراءات المتخذة',
            'severity_name': 'Severity',
            'stage_name': 'Stage',
            'harm_level': 'Harm',
            'status_name': 'Status',
            'feedback_risk_type': 'FeedbackRiskType',
            'section_answer': 'جواب القسم',
            'department_answer': 'جواب الدائرة',
            'administration_answer': 'جواب الإدارة',
            'last_edited': 'آخر تعديل',
        }
        
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = headers_arabic.get(col_name, col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Format dates
                if isinstance(value, (datetime, date)):
                    cell.value = value.strftime('%Y-%m-%d') if isinstance(value, date) else value.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    cell.value = value
                
                # Center align both horizontally and vertically for better readability
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, start=1):
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            if col_name in ['complaint_text', 'immediate_action', 'taken_action', 'section_answer', 'department_answer', 'administration_answer']:
                ws.column_dimensions[column_letter].width = 50
            elif col_name in ['patient_name', 'issuing_org_unit_name', 'concerned_org_unit_name', 'classification_name']:
                ws.column_dimensions[column_letter].width = 25
            elif col_name in ['incident_number', 'complaint_number']:
                ws.column_dimensions[column_letter].width = 18
            elif col_name in ['domain_name', 'category_name', 'subcategory_name']:
                ws.column_dimensions[column_letter].width = 20
            else:
                ws.column_dimensions[column_letter].width = 18
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
        
    finally:
        cursor.close()
        conn.close()


def export_complaints(
    export_format: Literal['csv', 'json'],
    filters: Dict[str, Any],
    columns: List[str],
    include_patient_identifiers: bool = False,
    language: str = 'en',
    restrict_to_unit_ids: Optional[set] = None,
) -> Dict[str, Any]:
    """
    Export filtered complaints as CSV or JSON.

    NOTE: This function prepares export metadata. Actual file generation
    should be handled by a background job or streaming response.
    """
    if export_format not in ['csv', 'json']:
        raise ValueError("Format must be 'csv' or 'json'")

    # Get count of records to export (scoped to user's org unit)
    count_result = get_complaints_count(**filters, restrict_to_unit_ids=restrict_to_unit_ids)
    record_count = count_result['total_count']
    
    # Generate export ID (timestamp-based)
    from datetime import datetime
    export_id = f"exp-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    
    # Determine file name
    start_date = filters.get('start_date', '')
    end_date = filters.get('end_date', '')
    if start_date and end_date:
        file_name = f"Complaints_Export_{start_date}_to_{end_date}.{export_format}"
    else:
        file_name = f"Complaints_Export_{datetime.now().strftime('%Y%m%d')}.{export_format}"
    
    return {
        'export_id': export_id,
        'file_name': file_name,
        'file_size_bytes': record_count * 500,  # Rough estimate
        'download_url': f"/api/complaints/download/{export_id}",
        'record_count': record_count,
        'generated_at': datetime.now().isoformat(),
        'expires_at': (datetime.now() + relativedelta(days=1)).isoformat(),
        'audit_logged': True,
        'status': 'pending'
    }


def get_cases_for_incident(incident_id: int) -> List[Dict[str, Any]]:
    """
    Get all cases for an incident with full field data (same shape as get_complaint_by_id).
    Used by EditRecord to load the complete incident with all its cases.
    """
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "SELECT IncidentRequestCaseID FROM dbo.APP_IncidentCase WHERE incident_id = ? ORDER BY IncidentRequestCaseID",
            (incident_id,)
        )
        case_ids = [row[0] for row in cursor.fetchall()]
    finally:
        conn.close()

    return [get_complaint_by_id(cid) for cid in case_ids]


def get_table_views() -> Dict[str, Any]:
    """
    Get predefined table view configurations.
    
    Returns:
        Dictionary with views array and default_view.
    """
    views = [
        {
            'view_id': 'complete',
            'view_name': 'Complete View',
            'view_name_ar': 'عرض كامل',
            'columns': [
                'complaint_number',
                'received_date',
                'patient_name',
                'issuing_org_unit_name',
                'domain_name',
                'category_name',
                'severity_name',
                'harm_level',
                'status_name',
                'days_open'
            ],
            'default_sort': 'received_date',
            'default_sort_order': 'desc'
        },
        {
            'view_id': 'simplified',
            'view_name': 'Simplified View',
            'view_name_ar': 'عرض مبسط',
            'columns': [
                'complaint_number',
                'received_date',
                'issuing_org_unit_name',
                'domain_name',
                'severity_name',
                'status_name'
            ],
            'default_sort': 'received_date',
            'default_sort_order': 'desc'
        }
    ]
    
    return {
        'views': views,
        'default_view': 'complete'
    }

