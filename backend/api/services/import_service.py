"""
Import Service — Hospital Data Intake Pipeline
Handles template generation, Excel parsing, validation, and controlled import.
"""

import base64
import hashlib
from datetime import datetime, date
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from core.database import get_connection
from ..db_layer import import_db
from ..db_layer import ml_import_batch_db
from ..db_layer.incident_parent import assign_case_to_incident
from .case_service import create_case
from . import staff_directory_service
from . import patient_directory_service

# Source-system tag for ml.ImportSourceRecordMap — generalizes the exact
# idempotency pattern already proven by dbo.APP_DataMigration_Map for the
# Phase K legacy-migration path (see ML_ARCHITECTURE_DECISION_RECORD.md 4.8).
EXCEL_IMPORT_SOURCE_SYSTEM = "ExcelImportTemplate"

# Injected into each parsed row dict to carry its real Excel sheet row
# number through grouping/validation -- lets the review screen reference a
# specific physical row for inline edits (see the /rows/{row_number} patch
# endpoints). Not a real template header, so it can't collide with one.
ROW_NUMBER_KEY = "__row_number__"

# Grouping key used when the Incident Number (Old System) is blank -- a real, presentable
# string (not an internal-looking sentinel) since it's shown directly in
# the review screen as the group's label.
MISSING_GROUP_KEY_LABEL = "(No Group Key)"

# Clinical risk type ID meaning "Ordinary" (no red flag / never event) —
# used as the default when the optional "Feedback Risk Type" column is blank,
# since case_service.create_case() requires clinical_risk_type_id.
DEFAULT_CLINICAL_RISK_TYPE_ID = 1

# Must match case_service.py's is_red_flag/is_never_event checks
# (clinical_risk_type_id == 2 / == 3) -- used here only to badge a row for
# human visibility in the review screen/report, since 'import_closed' save
# mode keeps every imported case Closed regardless of risk type.
RED_FLAG_RISK_TYPE_ID = 2
NEVER_EVENT_RISK_TYPE_ID = 3

# Where staged (uploaded-but-not-yet-confirmed) files live between
# stage_upload() and confirm_import(), keyed by ImportBatchID. Filesystem-
# based rather than a new DB column/table: no schema change needed, and it
# works fine across multiple backend worker processes on one server since
# they share a filesystem.
STAGING_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "import_staging"


def _staged_file_path(import_batch_id: int) -> Path:
    return STAGING_DIR / f"{import_batch_id}.xlsx"


# Generated batch reports are persisted here so the batch history page can
# offer a "download report" link for past batches, not just the one just
# confirmed -- the original uploaded rows for rejected groups aren't stored
# anywhere else (only accepted rows become real case records), so the
# report itself is the only durable record of what a past batch contained.
REPORTS_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "import_reports"


def _report_file_path(import_batch_id: int) -> Path:
    return REPORTS_DIR / f"{import_batch_id}.xlsx"


def get_report_path(import_batch_id: int) -> Optional[Path]:
    """Path to a past batch's saved report, or None if it doesn't exist."""
    path = _report_file_path(import_batch_id)
    return path if path.exists() else None


# ============================================================
# TEMPLATE COLUMN DEFINITIONS
# ============================================================

# (excel_header, lookup_list_key_or_inline_list, is_mandatory, col_width)
# Record Type is optional (defaults to Complaint if blank) rather than
# mandatory, so templates/sheets filled out before this column existed
# still work without every row being rejected.
# Feedback Type was dropped: dbo.APP_LOOKUP_FEEDBACK_INTENT_TYPE only has 2
# values ("Improvement Opportunity" / "Notice") that map 1:1 onto Record
# Type (Complaint/Notice) -- it's auto-derived instead, see
# _feedback_intent_type_id_for_record_type().
# Column order is deliberately NOT grouped by topic. It's split into 3 zones
# so the middle zone is a 1:1 positional mirror of the hospital's real source
# spreadsheet (Desktop/data.xlsx, sheet "أساسية(1+2+3)2025+1-2026") -- since a
# raw Excel block-paste is purely positional (it doesn't match by header
# text the way _parse_excel does), matching the source's own column order is
# what lets the team paste their whole sheet in one shot with zero manual
# reordering. Zone 1 (no source equivalent, entered manually, kept leftmost
# for visibility) -> Zone 2 (source mirror, placeholders inline for source
# columns we don't use, so nothing needs deleting/skipping before pasting)
# -> Zone 3 (remaining our-system-only fields, filled manually/left blank).
# Placeholder columns (lookup_key=None, mandatory=False, no COL_* constant
# below) are never read by _get()/_validate_group() -- they only exist to
# hold pasted source data so the block's column count/order lines up.
TEMPLATE_COLUMNS = [
    # ---- Zone 1: no source equivalent, manual entry ----
    ("Incident Date",               None,               True,  22),

    # ---- Zone 2: mirrors data.xlsx column-for-column ----
    ("Received Date",               None,               True,  22),
    # This IS the old system's Incident number -- not a meaningless legacy
    # ID. This system's Incident->Case structure was deliberately built to
    # mirror the old system's, so rows sharing this number are the same
    # incident there and must be grouped the same way here. _group_rows()
    # groups by this column (see COL_INCIDENT_KEY) instead of a separate
    # hand-typed key, which doesn't scale past a handful of rows -- typing a
    # correct, consistent grouping number for ~100 rows by hand is exactly
    # how rows for different patients ended up sharing one group number.
    ("Incident Number (Old System) (الرقم)", None,      True,  20),
    ("Patient Name",                None,               True,  25),
    ("Issuing Dept (قسم الصادر)",   "org_units",        True,  28),
    ("Administration (ادارة) — reference only, not imported", None, False, 24),
    ("Directorate Concerned (الدائرة المعنية) — reference only, not imported", None, False, 28),
    ("Target Dept",                 "org_units",        True,  28),
    ("Source (المصدر)",             "sources",          True,  20),
    ("Record Type",                 ["Complaint", "Notice"], False, 16),
    # Domain/Category/Subcategory are informational only -- the system
    # always derives these from the matched Classification
    # (classification_chains[class_id]); nothing typed in these 3 cells is
    # ever read. Labeled and marked non-mandatory accordingly, rather than
    # implying an independent requirement the backend doesn't enforce.
    ("Domain (auto-derived from Classification)", None, False, 20),
    ("Category (auto-derived from Classification)", None, False, 22),
    ("Subcategory (auto-derived from Classification)", None, False, 22),
    ("Classification (Arabic)",     "classifications",  True,  30),
    ("Classification (English)",    None,               False, 30),
    ("Complaint Text",              None,               True,  60),
    ("Immediate Action",            None,               False, 50),
    ("Taken Action (الإجراءات المتخذة)", None,          False, 50),
    ("Severity",                    "severities",       True,  16),
    ("Stage",                       "stages",           True,  16),
    ("Harm Level",                  "harm_levels",      True,  16),
    ("Status — reference only, not imported (imports are always closed)", None, False, 16),
    ("Feedback Risk Type",          "risk_types",       True,  20),

    # ---- Zone 3: no source equivalent, manual entry / optional ----
    ("Building",                    "buildings",        True,  18),
    ("Is Inpatient",                ["Yes", "No"],      False, 14),
    ("Doctor 1",                    "doctors",          False, 25),
    ("Doctor 2",                    "doctors",          False, 25),
    ("Doctor 3",                    "doctors",          False, 25),
    ("Worker 1 (Full Name)",        "workers",          False, 25),
    ("Worker 2 (Full Name)",        "workers",          False, 25),
    ("Worker 3 (Full Name)",        "workers",          False, 25),
]

# Column index constants (0-based). Only fields actually looked up by
# business logic get a constant -- the placeholder columns (including
# Domain/Category/Subcategory, which are display-only/derived) don't need
# one since nothing ever reads them.
COL_INCIDENT_DATE  = 0
COL_DATE           = 1  # Received Date
COL_INCIDENT_KEY   = 2  # Incident Number (Old System) (الرقم) -- the grouping key, see TEMPLATE_COLUMNS comment
COL_PATIENT        = 3
COL_ISSUING_DEPT   = 4
COL_TARGET_DEPT    = 7
COL_SOURCE         = 8
COL_RECORD_TYPE    = 9
COL_CLASS_AR       = 13
COL_CLASS_EN       = 14
COL_COMPLAINT      = 15
COL_IMMEDIATE      = 16
COL_TAKEN          = 17
COL_SEVERITY       = 18
COL_STAGE          = 19
COL_HARM           = 20
COL_RISK           = 22
COL_BUILDING       = 23
COL_INPATIENT      = 24
COL_DOCTOR1        = 25
COL_DOCTOR2        = 26
COL_DOCTOR3        = 27
COL_WORKER1        = 28
COL_WORKER2        = 29
COL_WORKER3        = 30

MAX_DATA_ROWS = 5000  # data validation applies up to this row; generous headroom so a slightly-oversized paste doesn't spill past the unlocked range

# Row layout of the "Import Template" sheet -- a merged zone-label row sits
# above the real header row (see generate_template's zone_bounds). Module-
# level (not just local to generate_template) because _validate_headers()
# and _parse_excel() below have to read from the SAME rows the template was
# generated with, or every upload looks like it's missing every column.
GROUP_ROW = 1
HEADER_ROW = 2
FIRST_DATA_ROW = 3
LAST_DATA_ROW = MAX_DATA_ROWS + 1

RECORD_TYPE_IDS = {
    "complaint": 1,
    "notice": 2,
    # The real source's "النوع" (Type) column holds this exact value on every
    # single row (verified: all 464 real rows) -- it's the hospital's generic
    # "improvement opportunity" tag, not a Complaint/Notice signal, so it
    # never varies and never actually discriminates between the two. Treated
    # as Complaint since that's what this whole feed is, and it's already
    # Record Type's own default when the field is blank -- without this
    # alias, every row hard-errors as "not recognized" instead.
    "فرصة تحسين": 1,
}
DEFAULT_RECORD_TYPE_ID = 1  # Complaint

# dbo.APP_LOOKUP_FEEDBACK_INTENT_TYPE: 1=Improvement Opportunity (used for
# Complaint records), 2=Notice (used for Notice records) -- confirmed
# 1:1 with Record Type, see TEMPLATE_COLUMNS comment above.
FEEDBACK_INTENT_TYPE_BY_RECORD_TYPE = {1: 1, 2: 2}


def _feedback_intent_type_id_for_record_type(record_type_id: int) -> int:
    return FEEDBACK_INTENT_TYPE_BY_RECORD_TYPE.get(record_type_id, 1)

DIRECTORY_LOOKUP_LIMIT = 500  # matches the Hospital Directory API's per-call max (see hospital_directory_client)


# ============================================================
# DIRECTORY LOOKUPS (doctors / workers / patients)
# ============================================================

def _load_directory_lookups() -> Dict[str, Dict[str, Any]]:
    """
    Merged doctor/worker lookups (reserve + Hospital Directory API),
    replacing the old APP_LOOKUP_DOCTOR/HR_EMPLOYEES_TABLE-only source.
    IDs may be a plain reserve int or an opaque external id string --
    case_service.create_case() already resolves either one via
    materialize_doctor_id()/materialize_employee_id(), so no extra
    resolution is needed here or at commit time.

    Capped at DIRECTORY_LOOKUP_LIMIT reserve + DIRECTORY_LOOKUP_LIMIT
    external entries per category (a single API page) -- a hospital with
    more active doctors/workers than that would need pagination added
    here; not attempted since it's well past what a dropdown list is
    usable for anyway.
    """
    maps: Dict[str, Any] = {}
    lists: Dict[str, Any] = {}

    doc_result = staff_directory_service.search_doctors_merged("", limit=DIRECTORY_LOOKUP_LIMIT)
    doctors = doc_result.get("doctors", []) if doc_result.get("success") else []
    maps["doctors"] = {(d["name"] or "").lower().strip(): d["doctor_id"] for d in doctors if d.get("name")}
    lists["doctors"] = [d["name"] for d in doctors if d.get("name")]

    wrk_result = staff_directory_service.search_workers_merged("", limit=DIRECTORY_LOOKUP_LIMIT)
    workers = wrk_result.get("employees", []) if wrk_result.get("success") else []
    maps["workers"] = {(w["full_name"] or "").lower().strip(): w["employee_id"] for w in workers if w.get("full_name")}
    lists["workers"] = [w["full_name"] for w in workers if w.get("full_name")]

    return {"maps": maps, "lists": lists}


def _load_lookups() -> Dict[str, Dict[str, Any]]:
    """import_db.load_all_lookups() plus the directory-sourced doctor/worker lists."""
    data = import_db.load_all_lookups()
    directory_data = _load_directory_lookups()
    data["maps"].update(directory_data["maps"])
    data["lists"].update(directory_data["lists"])
    return data


def _count_patient_matches(full_name: str) -> int:
    """
    Exact-name match count across reserve + Hospital Directory patients.
    Replaces the old count against the retired dbo.VW_PatientAdmission view.
    """
    result = patient_directory_service.search_patients_insert_flow(full_name, limit=50)
    if not result.get("success"):
        raise Exception(result.get("error") or "Patient search failed")
    target = full_name.strip().lower()
    return sum(
        1 for p in result["patients"]
        if (p.get("full_name") or "").strip().lower() == target
    )


# ============================================================
# TEMPLATE GENERATOR
# ============================================================

_INSTRUCTIONS_TITLE = "تعليمات استخدام قالب استيراد الحالات"

# Case, Incident, and Closed are kept in English inside the Arabic text on
# purpose -- staff were trained on the app using these English terms, so
# translating them would introduce unfamiliar vocabulary instead of helping.
_INSTRUCTIONS_TEXT = [
    "كل صف في هذا الملف يمثل Case واحد وليس Incident. الصفوف التي لها نفس "
    "رقم \"الرقم\" (Incident Number في النظام القديم) سيتم دمجها معًا تحت "
    "Incident واحد في النظام، تمامًا كما كانت في النظام القديم.",
    "عمود \"الرقم\" يجب أن يكون رقمًا فقط، بدون حروف أو رموز، ويجب تعبئته "
    "في كل صف حتى يتم تجميع الصفوف بشكل صحيح.",
    "الأعمدة التي تحتوي على قائمة منسدلة (مثل التصنيف، القسم، الطبيب) يجب اختيار "
    "القيمة منها فقط. لا تكتب نصًا غير موجود في القائمة، وإلا سيتم رفض الصف مع توضيح السبب.",
    "لا تقم بتغيير ترتيب الأعمدة أو أسماء العناوين أو حذف/إضافة أعمدة. "
    "الملف غير محمي تقنيًا، لكن أي تغيير في أسماء الأعمدة سيؤدي إلى رفض الملف "
    "بالكامل عند الرفع مع رسالة توضح السبب.",
    "إذا مر وقت طويل منذ تحميل هذا الملف، يفضل تحميل نسخة جديدة قبل التعبئة "
    "لضمان أن القوائم المنسدلة (الأطباء، الأقسام، وغيرها) محدثة.",
    "بعد رفع الملف، سيظهر لك تقرير يوضح كل Case تمت إضافتها بنجاح (مع رقمها "
    "الجديد في النظام) وكل Case تم رفضها مع سبب الرفض بوضوح.",
    "جميع الـ Cases التي يتم استيرادها تدخل النظام بحالة Closed مباشرة، "
    "ولا تمر عبر نظام الرسائل أو صندوق الوارد.",
    "إذا كان الطبيب أو الموظف غير موجود في القائمة، يمكنك ترك الخانة فارغة "
    "والمتابعة، أو إضافته أولاً من صفحة الإعدادات > الأطباء ثم إعادة تحميل القالب.",
]


def _add_instructions_sheet(wb: Workbook) -> None:
    """
    Visible, RTL instructions sheet, first tab in the file (but not the
    active one on open -- see generate_template's wb.active -- so the file
    opens straight onto the fillable grid; Instructions is one click away).
    Numbered-card layout: badge column + wrapped text column, bordered rows.
    """
    ins = wb.create_sheet("تعليمات - Instructions", 0)
    ins.sheet_view.rightToLeft = True
    ins.sheet_view.showGridLines = False
    ins.column_dimensions["A"].width = 6
    ins.column_dimensions["B"].width = 100

    title_fill = PatternFill("solid", fgColor="1F4E79")
    badge_fill = PatternFill("solid", fgColor="2E75B6")
    card_fill = PatternFill("solid", fgColor="F2F7FC")
    thin = Side(style="thin", color="D0DCE8")
    card_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title banner
    ins.merge_cells("A1:B1")
    title_cell = ins.cell(row=1, column=1, value=_INSTRUCTIONS_TITLE)
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ins.row_dimensions[1].height = 36

    row_i = 3
    for i, line in enumerate(_INSTRUCTIONS_TEXT, start=1):
        badge = ins.cell(row=row_i, column=1, value=i)
        badge.font = Font(bold=True, size=13, color="FFFFFF")
        badge.fill = badge_fill
        badge.alignment = Alignment(horizontal="center", vertical="center")
        badge.border = card_border

        text = ins.cell(row=row_i, column=2, value=line)
        text.font = Font(size=12)
        text.fill = card_fill
        text.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        text.border = card_border

        # Rough auto-height: ~45 chars per line at this column width/font size
        ins.row_dimensions[row_i].height = max(30, 15 * (len(line) // 45 + 1))
        ins.row_dimensions[row_i + 1].height = 8  # spacer between cards
        row_i += 2

    ins.protection.sheet = True
    ins.protection.formatCells = True
    ins.protection.formatColumns = True
    ins.protection.insertColumns = True
    ins.protection.deleteColumns = True


def generate_template() -> BytesIO:
    """
    Build the Excel import template with live DB dropdowns.
    Returns BytesIO of the .xlsx file.
    """
    data = _load_lookups()
    lookup_lists = data["lists"]

    wb = Workbook()

    # ---- Hidden lookups sheet ----
    ls = wb.active
    ls.title = "Lookups"
    ls.sheet_state = "hidden"

    # Each lookup category occupies one column in the Lookups sheet
    # Col order matches the TEMPLATE_COLUMNS lookup keys
    lookup_key_to_col: Dict[str, int] = {}   # key -> 1-based col number in Lookups sheet
    lookup_col = 1
    for _, lookup_key, _, _ in TEMPLATE_COLUMNS:
        if lookup_key is None or isinstance(lookup_key, list):
            continue
        if lookup_key in lookup_key_to_col:
            continue
        lookup_key_to_col[lookup_key] = lookup_col
        ls.cell(row=1, column=lookup_col, value=lookup_key)  # header row
        values = lookup_lists.get(lookup_key, [])
        for row_i, val in enumerate(values, start=2):
            ls.cell(row=row_i, column=lookup_col, value=val)
        lookup_col += 1

    # ---- Template sheet ----
    ts = wb.create_sheet("Import Template", 1)
    ts.sheet_view.rightToLeft = True

    # Row layout (GROUP_ROW/HEADER_ROW/FIRST_DATA_ROW/LAST_DATA_ROW) is
    # module-level now -- see the comment near MAX_DATA_ROWS. _validate_headers()
    # and _parse_excel() read from the same constants, so the two stay in sync.
    header_fill = PatternFill("solid", fgColor="1F4E79")
    mandatory_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col_i, (header, lookup_key, mandatory, width) in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ts.cell(row=HEADER_ROW, column=col_i, value=header)
        cell.fill = mandatory_fill if mandatory else header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ts.column_dimensions[get_column_letter(col_i)].width = width

    ts.row_dimensions[HEADER_ROW].height = 40

    # ---- Zone-label row: groups the columns into 3 visual bands so the
    # team can see at a glance which block to paste their Excel data into.
    # Derived from the COL_* constants (not hardcoded numbers) so it stays
    # correct if TEMPLATE_COLUMNS' zone boundaries ever move -- each zone is
    # assumed to start at the constant named here (true by construction,
    # see the comment above TEMPLATE_COLUMNS).
    zone_bounds = [
        (COL_INCIDENT_DATE + 1, COL_DATE,  "1F4E5A", "PART 1 — Fill In Manually"),
        (COL_DATE + 1,      COL_BUILDING,  "C0522B", "PART 2 — Paste Your Excel Data Here"),
        (COL_BUILDING + 1,  len(TEMPLATE_COLUMNS), "5B6B73", "PART 3 — Additional Info (Manual / Optional)"),
    ]
    for start_col, end_col, color, label in zone_bounds:
        ts.merge_cells(start_row=GROUP_ROW, start_column=start_col, end_row=GROUP_ROW, end_column=end_col)
        cell = ts.cell(row=GROUP_ROW, column=start_col, value=label)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # merge_cells only styles the top-left cell -- apply fill/border to
        # every cell in the merged range so the band of color is continuous.
        for col_i in range(start_col, end_col + 1):
            ts.cell(row=GROUP_ROW, column=col_i).fill = PatternFill("solid", fgColor=color)
            ts.cell(row=GROUP_ROW, column=col_i).border = border
    ts.row_dimensions[GROUP_ROW].height = 40

    ts.freeze_panes = f"A{FIRST_DATA_ROW}"

    # Apply data validation dropdowns (+ a couple of special-cased columns
    # that aren't list-based lookups: date type-checking, numeric-only Incident Number)
    for col_i, (_, lookup_key, _, _) in enumerate(TEMPLATE_COLUMNS, start=1):
        col_letter = get_column_letter(col_i)
        cell_range = f"{col_letter}{FIRST_DATA_ROW}:{col_letter}{LAST_DATA_ROW}"
        col_index0 = col_i - 1

        if col_index0 in (COL_INCIDENT_DATE, COL_DATE):
            dv = DataValidation(
                type="date", operator="greaterThan", formula1="1900-01-01",
                allow_blank=True,
            )
            dv.error = "Please enter a valid date (use the cell's date picker)."
            dv.errorTitle = "Invalid date"
            dv.sqref = cell_range
            ts.add_data_validation(dv)
            for row_i in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
                ts.cell(row=row_i, column=col_i).number_format = "yyyy-mm-dd"
            continue

        if col_index0 == COL_INCIDENT_KEY:
            dv = DataValidation(
                type="whole", operator="greaterThan", formula1="0",
                allow_blank=True,
            )
            dv.error = "Incident Number (Old System) must be a number only, no letters."
            dv.errorTitle = "Numbers only"
            dv.sqref = cell_range
            ts.add_data_validation(dv)
            continue

        if lookup_key is None:
            continue

        if isinstance(lookup_key, list):
            # Inline small list (e.g. Yes/No)
            formula = '"' + ",".join(lookup_key) + '"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        else:
            ls_col = lookup_key_to_col.get(lookup_key)
            if ls_col is None:
                continue
            values = lookup_lists.get(lookup_key, [])
            if not values:
                continue
            ls_col_letter = get_column_letter(ls_col)
            last_row = len(values) + 1
            formula = f"'Lookups'!${ls_col_letter}$2:${ls_col_letter}${last_row}"
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)

        dv.sqref = cell_range
        ts.add_data_validation(dv)

    # Instructions row (FIRST_DATA_ROW is the first real data row)
    ts.cell(row=FIRST_DATA_ROW, column=COL_PATIENT + 1).value = "Patient Name Here"
    # Incident Date deliberately gets NO example value (same treatment as
    # Target Dept/Doctor/Worker below) -- it has no source equivalent and
    # must be filled manually per row. A real, valid-looking date here was a
    # silent trap: if a user builds many rows by extending/dragging from the
    # example row instead of retyping Incident Date on every single one,
    # today's date rode along on every row while the correctly-pasted
    # Received Date showed the real historical date, tripping the (correctly
    # working) "Incident Date cannot be after Received Date" check.
    ts.cell(row=FIRST_DATA_ROW, column=COL_INCIDENT_DATE + 1).number_format = "yyyy-mm-dd"
    ts.cell(row=FIRST_DATA_ROW, column=COL_DATE + 1).value = datetime.today().date()
    ts.cell(row=FIRST_DATA_ROW, column=COL_DATE + 1).number_format = "yyyy-mm-dd"
    ts.cell(row=FIRST_DATA_ROW, column=COL_COMPLAINT + 1).value = "Complaint text here"

    # Style the example row lightly
    example_fill = PatternFill("solid", fgColor="EBF3FB")
    for col_i in range(1, len(TEMPLATE_COLUMNS) + 1):
        cell = ts.cell(row=FIRST_DATA_ROW, column=col_i)
        cell.fill = example_fill

    # No sheet protection at all -- column identity (the original reason this
    # sheet was ever locked down) is enforced server-side instead, by name,
    # at upload time (see _validate_headers), which is a real guarantee
    # regardless of any client-side Excel lock (there was never a password on
    # it anyway, so it was never a hard barrier). Every restriction protection
    # used to add here -- paste, resize, insert, sort, and even just clicking
    # near the header/zone-label rows -- was recurring real friction for
    # near-zero actual safety benefit, so it's gone. If a header ever does
    # get renamed/deleted, the failure mode is a clear upload-time rejection
    # with instructions to re-download, not silent data corruption.
    #
    # The long free-text columns (Complaint Text, Immediate Action, Taken
    # Action) still get wrap_text so paragraph-length values stay readable
    # instead of clipping, matching how the source spreadsheet itself
    # displays them (tall wrapped rows). No explicit row height is set on
    # purpose -- Excel auto-fits row height to wrapped content as long as the
    # row's height isn't pinned to a fixed value, which none of these are.
    wide_text_cols = {COL_COMPLAINT, COL_IMMEDIATE, COL_TAKEN}
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for row in ts.iter_rows(min_row=FIRST_DATA_ROW, max_row=LAST_DATA_ROW, max_col=len(TEMPLATE_COLUMNS)):
        for cell in row:
            if cell.column - 1 in wide_text_cols:
                cell.alignment = wrap_alignment

    _add_instructions_sheet(wb)
    # Instructions is the first tab (visible, one click away) but the file
    # should open straight onto the fillable grid, not a wall of text.
    wb.active = wb.sheetnames.index(ts.title)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# UPLOAD PIPELINE
# ============================================================

def _is_flagged_risk(risk_type_id: Optional[int]) -> bool:
    """Red flag or never event — see RED_FLAG_RISK_TYPE_ID/NEVER_EVENT_RISK_TYPE_ID."""
    return risk_type_id in (RED_FLAG_RISK_TYPE_ID, NEVER_EVENT_RISK_TYPE_ID)


def _summarize_row_errors(row_results: List[Dict[str, Any]]) -> str:
    """One-line summary of a group's per-row errors, for contexts that only
    show a single reason string (the report Excel, legacy display)."""
    parts = []
    for r in row_results:
        if r["errors"]:
            messages = "; ".join(e["message"] for e in r["errors"])
            parts.append(f"Row {r['row_number']}: {messages}")
    return " | ".join(parts) if parts else "Unknown validation error"


def _validate_and_group(file_bytes: bytes) -> Dict[str, Any]:
    """
    Parse -> group by Incident Number (Old System) -> validate each group -> check
    per-group duplicate status. Pure computation, no case/incident writes —
    shared by stage_upload() (preview) and confirm_import() (re-validated
    fresh right before commit, so anything that changed in the lookups
    between preview and confirm — e.g. a doctor added in Settings — is
    picked up automatically).
    """
    data = _load_lookups()
    maps = data["maps"]

    rows = _parse_excel(file_bytes)

    groups = _group_rows(rows)
    valid_groups: List[Tuple[str, List[dict], dict]] = []
    rejected_groups: List[Dict[str, Any]] = []
    all_warnings: List[Dict[str, str]] = []

    for group_key, group_rows in groups.items():
        result = _validate_group(group_key, group_rows, maps)
        if result["valid"]:
            valid_groups.append((group_key, group_rows, result))
        else:
            rejected_groups.append({
                "group_key": group_key,
                "rows": group_rows,
                "reason": _summarize_row_errors(result["rows"]),
                "row_results": result["rows"],
                "status": "red",
            })
        all_warnings.extend(result.get("warnings", []))

    # NOTE: there is deliberately no cross-upload duplicate check on Incident
    # Group Key here. The key is purely a within-file grouping label (rows
    # sharing the same key become one incident) -- it resets to 1, 2, 3... on
    # every file, so the same key is expected to reappear in every upload.
    # Treating it as a durable external identifier (which an earlier version
    # of this pipeline did, via find_group_already_imported) meant the very
    # first import using key "1" permanently blocked every future import that
    # also used "1", i.e. almost all of them. Batch-level dedup (the exact
    # same file re-uploaded, checked via file_checksum in stage_upload) is
    # the real duplicate guard; _import_group's external_record_id is scoped
    # per-batch so it stays unique without reintroducing that check.

    return {
        "rows": rows,
        "groups": groups,
        "valid_groups": valid_groups,
        "rejected_groups": rejected_groups,
        "warnings": all_warnings,
    }


def _json_safe(value: Any) -> Any:
    """Make a raw cell value JSON-serializable (dates/datetimes -> ISO strings)."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _row_preview(row_result: Dict[str, Any]) -> Dict[str, Any]:
    """
    Shape one row for the review grid using the CURRENT template's column
    set, not whatever headers happened to be in the uploaded file. A file
    uploaded before a template change (e.g. missing the newer Incident Date
    column, or still carrying a since-removed one like the old Feedback
    Type) would otherwise leak stale, unpatchable field names into the grid
    -- editing them would silently do nothing, since patch_staged_rows()
    only recognizes today's TEMPLATE_COLUMNS headers. Columns the file
    never had simply show up blank here, ready to be filled in through the
    grid even though the original upload didn't have them.
    """
    raw = row_result["raw"]
    return {
        "row_number": row_result["row_number"],
        "errors": row_result["errors"],
        "fields": {col[0]: _json_safe(raw.get(col[0])) for col in TEMPLATE_COLUMNS},
        "derived_hierarchy": row_result.get("derived_hierarchy") or {"domain": None, "category": None, "subcategory": None},
        "resolved_display": row_result.get("resolved_display") or {"record_type": None},
    }


def _build_preview_groups(valid_groups: List[Tuple[str, List[dict], dict]],
                           rejected_groups: List[Dict[str, Any]],
                           order: Optional[List[str]] = None) -> List[Dict[str, Any]]:
    """
    Shape validated/rejected groups for the review screen: one entry per
    Incident Number (Old System), colored per the decided taxonomy (green/yellow ready,
    red/duplicate blocked), with a red-flag/never-event badge that's purely
    informational -- it never blocks the group or changes its color. Each
    group carries its individual rows (row_number, per-row errors, editable
    field values) so the review grid can show and fix a specific row.

    `order` is the Incident Number (Old System) sequence as it first appeared in the
    uploaded file (see _group_rows -- a plain dict, insertion-ordered). Valid
    and rejected groups are validated as two separate passes below, which
    would otherwise bunch every ready group first and every rejected group
    after; sorting by `order` at the end restores the file's natural
    row-by-row order so a reviewer scanning the screen sees groups in the
    same sequence they appear in Excel, statuses interleaved.
    """
    preview: List[Dict[str, Any]] = []

    for group_key, group_rows, validation in valid_groups:
        validated_rows = [r["data"] for r in validation["rows"]]
        preview.append({
            "group_key": group_key,
            "status": "yellow" if (validation["is_new_patient"] or validation.get("has_fuzzy_match")) else "green",
            "is_new_patient": validation["is_new_patient"],
            "has_fuzzy_match": validation.get("has_fuzzy_match", False),
            "patient_name": validation["patient_name"],
            "row_count": len(validated_rows),
            "has_flagged_risk": any(_is_flagged_risk(r.get("risk_type_id")) for r in validated_rows),
            "reason": None,
            "rows": [_row_preview(r) for r in validation["rows"]],
        })

    for rejected in rejected_groups:
        preview.append({
            "group_key": rejected["group_key"],
            "status": rejected["status"],  # "red" or "duplicate"
            "is_new_patient": False,
            "patient_name": None,
            "row_count": len(rejected["rows"]),
            "has_flagged_risk": False,
            "reason": rejected["reason"],
            "rows": [_row_preview(r) for r in rejected.get("row_results", [])],
        })

    if order is not None:
        order_index = {key: i for i, key in enumerate(order)}
        preview.sort(key=lambda g: order_index.get(g["group_key"], len(order_index)))

    return preview


def stage_upload(file_bytes: bytes, uploaded_by_user_id: int = 1) -> Dict[str, Any]:
    """
    Phase 1: checksum dedup -> parse -> group -> validate -> per-group
    duplicate check -> build a review-screen preview grouped by incident.
    Nothing is written to APP_Incident/APP_IncidentCase here — the staged
    file is persisted so confirm_import(import_batch_id) can commit it later.
    """
    header_error = _validate_headers(file_bytes)
    if header_error:
        return _empty_report(header_error)

    file_checksum = hashlib.sha256(file_bytes).hexdigest()

    # Batch-level duplicate check — has this exact file already been
    # imported successfully? (see ML_ARCHITECTURE_DECISION_RECORD.md 4.8)
    batch_conn = get_connection()
    batch_cursor = batch_conn.cursor()
    try:
        existing_batch = ml_import_batch_db.find_batch_by_checksum(batch_cursor, file_checksum)
        if existing_batch:
            return _empty_report(
                f"This exact file was already imported successfully on "
                f"{existing_batch['UploadedAt']} (batch #{existing_batch['ImportBatchID']}). "
                f"Re-upload blocked to prevent duplicate cases."
            )

        import_batch_id = ml_import_batch_db.create_import_batch(
            batch_cursor,
            original_file_name=None,
            file_checksum=file_checksum,
            template_version="v1",
            uploaded_by_user_id=uploaded_by_user_id,
        )
        batch_conn.commit()
    finally:
        batch_cursor.close()
        batch_conn.close()

    STAGING_DIR.mkdir(parents=True, exist_ok=True)
    _staged_file_path(import_batch_id).write_bytes(file_bytes)

    validated = _validate_and_group(file_bytes)
    if not validated["rows"]:
        return _empty_report("No data rows found in the uploaded file.")

    preview_groups = _build_preview_groups(validated["valid_groups"], validated["rejected_groups"], order=list(validated["groups"].keys()))

    summary_conn = get_connection()
    summary_cursor = summary_conn.cursor()
    try:
        ml_import_batch_db.update_batch_summary(
            summary_cursor, import_batch_id,
            total_rows=len(validated["rows"]),
            status="PendingReview",
        )
        summary_conn.commit()
    finally:
        summary_cursor.close()
        summary_conn.close()

    return {
        "import_batch_id": import_batch_id,
        "status": "PendingReview",
        "total_rows": len(validated["rows"]),
        "groups": preview_groups,
        "warnings": validated["warnings"],
    }


def _load_pending_batch(import_batch_id: int) -> Dict[str, Any]:
    """Fetch a batch and assert it's still PendingReview, or raise ValueError."""
    conn = get_connection()
    cursor = conn.cursor()
    try:
        batch = ml_import_batch_db.get_batch(cursor, import_batch_id)
    finally:
        cursor.close()
        conn.close()

    if batch is None:
        raise ValueError(f"Import batch {import_batch_id} not found.")
    if batch["Status"] != "PendingReview":
        raise ValueError(
            f"Import batch {import_batch_id} is not awaiting review (status: {batch['Status']})."
        )
    return batch


def patch_staged_rows(import_batch_id: int, patches: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Apply field edits directly onto the staged Excel file for a still-
    PendingReview batch, then re-validate. patches: [{"row_number": int,
    "fields": {header: value}}, ...] -- the caller (PATCH /rows/{n}) always
    passes a single-row list, but this stays list-shaped since the file
    open/save is one cycle regardless of how many rows it covers. No other
    code path needs to change: resume_preview()/confirm_import() already
    just re-read whatever's currently in the staged file.

    Incident Number (Old System) -- the grouping key -- is deliberately not
    patchable here -- changing it would move a row into a different incident
    group entirely, which needs a full re-group, not a single-cell edit. Fix
    that one in Excel and re-upload, same as before this feature existed.
    """
    _load_pending_batch(import_batch_id)  # raises ValueError if not found/not PendingReview

    staged_path = _staged_file_path(import_batch_id)
    if not staged_path.exists():
        raise ValueError(
            f"Staged file for batch {import_batch_id} is no longer available — please re-upload."
        )

    header_to_col = {col[0]: i + 1 for i, col in enumerate(TEMPLATE_COLUMNS)}
    group_key_header = TEMPLATE_COLUMNS[COL_INCIDENT_KEY][0]
    date_headers = {TEMPLATE_COLUMNS[COL_INCIDENT_DATE][0], TEMPLATE_COLUMNS[COL_DATE][0]}

    wb = load_workbook(staged_path)
    ws = wb["Import Template"] if "Import Template" in wb.sheetnames else wb.active

    for patch in patches:
        row_number = patch.get("row_number")
        if not isinstance(row_number, int) or row_number < FIRST_DATA_ROW:
            continue
        for header, value in (patch.get("fields") or {}).items():
            if header == group_key_header:
                continue  # not patchable, see docstring
            col_i = header_to_col.get(header)
            if col_i is None:
                continue  # unknown field name, ignore rather than error the whole batch
            cell = ws.cell(row=row_number, column=col_i)
            if header in date_headers and value:
                cell.value = datetime.strptime(str(value), "%Y-%m-%d").date()
                cell.number_format = "yyyy-mm-dd"
            else:
                cell.value = value if value not in ("", None) else None

    wb.save(staged_path)

    validated = _validate_and_group(staged_path.read_bytes())
    preview_groups = _build_preview_groups(validated["valid_groups"], validated["rejected_groups"], order=list(validated["groups"].keys()))
    return {
        "import_batch_id": import_batch_id,
        "status": "PendingReview",
        "total_rows": len(validated["rows"]),
        "groups": preview_groups,
        "warnings": validated["warnings"],
    }


def get_editable_lookups() -> Dict[str, List[str]]:
    """
    Plain internal lookup lists for the review grid's lookup-field editors
    (Classification, Department, Domain, etc.) -- small, fetched once,
    filtered client-side. Doctor/Worker are deliberately excluded: those
    use live search instead (GET /api/records/search/doctors|employees),
    same directory the manual Insert Record form already searches.
    """
    data = _load_lookups()
    lists = dict(data["lists"])
    lists.pop("doctors", None)
    lists.pop("workers", None)
    return lists


def discard_batch(import_batch_id: int) -> None:
    """Discard a still-PendingReview batch: delete its DB row and staged file."""
    conn = get_connection()
    cursor = conn.cursor()
    try:
        batch = ml_import_batch_db.get_batch(cursor, import_batch_id)
        if batch is None:
            raise ValueError(f"Import batch {import_batch_id} not found.")
        if batch["Status"] == "Completed":
            raise ValueError(
                f"Import batch {import_batch_id} has already been confirmed and can't be discarded."
            )
        ml_import_batch_db.delete_batch(cursor, import_batch_id)
        conn.commit()
    finally:
        cursor.close()
        conn.close()

    staged_path = _staged_file_path(import_batch_id)
    if staged_path.exists():
        staged_path.unlink()


def resume_preview(import_batch_id: int) -> Dict[str, Any]:
    """
    Re-open the review screen for a batch that's still PendingReview --
    e.g. the user uploaded, then refreshed/navigated away before clicking
    Add, losing the in-browser preview state. Re-reads the still-staged
    file and re-validates fresh, same as stage_upload()'s first pass, just
    without re-creating the batch row or re-checking the file checksum.
    """
    _load_pending_batch(import_batch_id)

    staged_path = _staged_file_path(import_batch_id)
    if not staged_path.exists():
        raise ValueError(
            f"Staged file for batch {import_batch_id} is no longer available — please re-upload."
        )
    file_bytes = staged_path.read_bytes()

    validated = _validate_and_group(file_bytes)
    preview_groups = _build_preview_groups(validated["valid_groups"], validated["rejected_groups"], order=list(validated["groups"].keys()))

    return {
        "import_batch_id": import_batch_id,
        "status": "PendingReview",
        "total_rows": len(validated["rows"]),
        "groups": preview_groups,
        "warnings": validated["warnings"],
    }


def confirm_import(import_batch_id: int, confirmed_by_user_id: int = 1) -> Dict[str, Any]:
    """
    Phase 2: re-validate the staged file fresh (catches anything that
    changed in the lookups since staging) and actually commit every fully
    valid incident group. Call only after the user has reviewed
    stage_upload()'s preview and clicked Add.
    """
    _load_pending_batch(import_batch_id)

    staged_path = _staged_file_path(import_batch_id)
    if not staged_path.exists():
        raise ValueError(
            f"Staged file for batch {import_batch_id} is no longer available — please re-upload."
        )
    file_bytes = staged_path.read_bytes()

    validated = _validate_and_group(file_bytes)
    rows = validated["rows"]
    groups = validated["groups"]
    valid_groups = validated["valid_groups"]
    rejected_groups = validated["rejected_groups"]
    all_warnings = validated["warnings"]

    # Import valid groups
    imported: List[Dict[str, Any]] = []
    imported_with_rows: List[Tuple[Dict[str, Any], List[Dict]]] = []
    new_patients_count = 0

    for group_key, group_rows, validation in valid_groups:
        try:
            validated_rows = [r["data"] for r in validation["rows"]]
            imp_result = _import_group(group_key, validated_rows,
                                       validation["is_new_patient"], confirmed_by_user_id,
                                       import_batch_id)
            imp_result["has_flagged_risk"] = any(
                _is_flagged_risk(r.get("risk_type_id")) for r in validated_rows
            )
            imported.append(imp_result)
            imported_with_rows.append((imp_result, group_rows))
            if validation["is_new_patient"]:
                new_patients_count += 1
        except Exception as exc:
            rejected_groups.append({
                "group_key": group_key,
                "rows": group_rows,
                "reason": f"Import error: {exc}",
                "status": "red",
            })

    # Full report Excel: green (imported, with new Incident Number) +
    # red/blue (rejected/duplicate, with reason) -- always generated, not
    # just when something was rejected, since it's the receipt of the batch.
    report_buf = _generate_import_report_excel(imported_with_rows, rejected_groups)
    report_bytes = report_buf.getvalue()
    rejected_b64 = base64.b64encode(report_bytes).decode()

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    _report_file_path(import_batch_id).write_bytes(report_bytes)

    total_imported_rows = sum(r["rows_imported"] for r in imported)
    total_failed_rows_within_groups = sum(r.get("rows_failed", 0) for r in imported)
    total_rejected_rows = sum(len(r["rows"]) for r in rejected_groups) + total_failed_rows_within_groups

    # Update batch summary (best-effort — never blocks the response)
    try:
        summary_conn = get_connection()
        summary_cursor = summary_conn.cursor()
        try:
            ml_import_batch_db.update_batch_summary(
                summary_cursor, import_batch_id,
                total_rows=len(rows),
                accepted_rows=total_imported_rows,
                rejected_rows=total_rejected_rows,
                created_case_count=total_imported_rows,
                status="Completed",
            )
            summary_conn.commit()
        finally:
            summary_cursor.close()
            summary_conn.close()
    except Exception as e:
        print(f"[IMPORT BATCH WARNING] Failed to update batch summary: {e}")

    # The staged file has served its purpose — nothing left to re-confirm.
    try:
        staged_path.unlink(missing_ok=True)
    except OSError:
        pass

    return {
        "summary": {
            "import_batch_id": import_batch_id,
            "total_groups": len(groups),
            "imported_groups": len(imported),
            "rejected_groups": len(rejected_groups),
            "total_rows": len(rows),
            "imported_rows": total_imported_rows,
            "rejected_rows": total_rejected_rows,
            "new_patients_created": new_patients_count,
            "warnings_count": len(all_warnings),
        },
        "imported": imported,
        "rejected": [
            {"group_key": r["group_key"], "reason": r["reason"], "status": r.get("status", "red")}
            for r in rejected_groups
        ],
        "warnings": all_warnings,
        "rejected_excel_b64": rejected_b64,
    }


def process_upload(file_bytes: bytes, created_by_user_id: int = 1) -> Dict[str, Any]:
    """
    One-shot convenience wrapper: stage then immediately confirm, preserving
    the pre-review-screen behavior (used by existing tests and any caller
    that doesn't need the review step). New UI flows should call
    stage_upload() and confirm_import() separately instead.
    """
    staged = stage_upload(file_bytes, uploaded_by_user_id=created_by_user_id)
    if "import_batch_id" not in staged:
        return staged  # blocked before staging (e.g. duplicate file, empty file)
    return confirm_import(staged["import_batch_id"], confirmed_by_user_id=created_by_user_id)


# ============================================================
# INTERNAL — PARSE
# ============================================================

def _validate_headers(file_bytes: bytes) -> Optional[str]:
    """
    Confirm every expected column name is present somewhere in HEADER_ROW.
    _parse_excel below keys each row by whatever header TEXT actually sits
    in HEADER_ROW, not by column position, so a column dragged to a new
    position is harmless -- its header and data move together and the
    name-based lookup still finds it. A renamed or deleted header is the one
    edit that silently drops that column's data with no error, so that's
    what this catches, up front, before staging or parsing anything.
    Order-independent on purpose: reordering isn't a correctness problem,
    only identity is.
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = None
    for name in wb.sheetnames:
        if name == "Import Template":
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.active

    actual_headers = {str(c.value).strip() for c in sheet[HEADER_ROW] if c.value is not None}
    expected_headers = {col[0] for col in TEMPLATE_COLUMNS}
    missing = expected_headers - actual_headers
    if not missing:
        return None
    return (
        "This file is missing expected column(s): " + ", ".join(sorted(missing)) +
        ". A column header may have been renamed or deleted. Please re-download "
        "a fresh template and copy your data into it, keeping the original "
        "column headers intact."
    )


def _parse_excel(file_bytes: bytes) -> List[Dict[str, Any]]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    # Prefer "Import Template" sheet, else use first visible sheet
    sheet = None
    for name in wb.sheetnames:
        if name == "Import Template":
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.active

    headers = [cell.value for cell in sheet[HEADER_ROW]]
    rows = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=FIRST_DATA_ROW, values_only=True), start=FIRST_DATA_ROW):
        # Skip completely empty rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        row_dict = {ROW_NUMBER_KEY: row_number}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = val
        rows.append(row_dict)
    return rows


# ============================================================
# INTERNAL — GROUP
# ============================================================

def _group_rows(rows: List[Dict]) -> Dict[str, List[Dict]]:
    groups: Dict[str, List[Dict]] = {}
    header = TEMPLATE_COLUMNS[COL_INCIDENT_KEY][0]
    for row in rows:
        key_raw = row.get(header)
        if isinstance(key_raw, (int, float)) and float(key_raw).is_integer():
            # Excel hands back a real number for the numeric-only Incident
            # Number column -- normalize 1.0/1 to the same "1" grouping key.
            key = str(int(key_raw))
        else:
            key = str(key_raw).strip() if key_raw is not None else ""
        if not key:
            key = MISSING_GROUP_KEY_LABEL
        groups.setdefault(key, []).append(row)
    return groups


# ============================================================
# INTERNAL — VALIDATE
# ============================================================

def _get(row: Dict, col_index: int) -> Optional[str]:
    """Get cell value by column index, return stripped string or None."""
    header = TEMPLATE_COLUMNS[col_index][0]
    val = row.get(header)
    if val is None:
        return None
    return str(val).strip() or None


def _lookup(maps: Dict, category: str, value: Optional[str]) -> Optional[int]:
    if not value:
        return None
    return maps.get(category, {}).get(value.lower().strip())


FUZZY_MATCH_THRESHOLD = 0.87  # conservative; substring containment scores 0.99 and is checked first regardless

# Curated "bank" of known real-world variants per DB value, keyed by the
# DB value's lowercased form (matching maps[category]'s key space) -> list
# of additional lowercased phrasings actually seen in real hospital exports.
# Minimum 0 extra entries (the DB value itself is always an implicit
# candidate), no maximum -- extend this as new variants are spotted.
# Exists because some real variants share almost no characters with the
# canonical DB value (e.g. "Moderate" vs DB "Medium" scores 0.29 similarity
# -- nowhere near any workable threshold) and can only be caught by an
# explicit alias, not by character-sequence matching alone.
LOOKUP_ALIASES: Dict[str, Dict[str, List[str]]] = {
    "severities": {
        "medium": ["moderate"],
    },
    "harm_levels": {
        "moderate": ["moderate harm"],
        "minor": ["minor harm"],
        "severe": ["severe harm", "high severe"],
    },
    "risk_types": {
        "ordinary": ["ordinary complaint"],
    },
}

# Classification -> known English phrasing(s), keyed by the DB's Arabic
# value (the DB's `classifications` lookup is Arabic-only -- confirmed
# against the real DB list; English classification text scores 0.00-0.18
# similarity against it, a categorical script mismatch no threshold fixes).
# Built empirically from a real 464-row hospital export, where the source
# already pairs an Arabic and an English classification per row for the same
# underlying value: 75 of 77 distinct Arabic values paired 1:1 with a single
# English phrasing; 2 have more than one observed English phrasing due to
# real inconsistent data entry (kept as multiple entries here on purpose --
# _match_classification tries all of them and takes the best).
CLASSIFICATION_EN_ALIASES: Dict[str, List[str]] = {
    'خلل في متابعة حالة المريض': ['Examination/Monitoring Problems'],
    'تجاهل المريض': ['Ignoring Patients'],
    'خلل في العناية التمريضية(الحفاض,إجراء الحمام..)': ['Nursing Care Problems(Diapper,bath..)'],
    'نقص في مهارة التمريض(المصل..)': ['IV Problem(Nursing Skills..)'],
    'انتظار الأدوار': ['Delay Access(Waiting for Consultation..)'],
    'برتوكول طبي': ['Medical Protocol'],
    'مشاكل في المستلزمات (الشراشف/وسادة/حرام...)': ['Problems in ihe facilities(pillows, covers..)'],
    'مواعيد بعيدة(العيادات الخارجية,القلبية..)': ['Delay Access(Clinic Appointment)'],
    'خلل في إجراءات حماية المريض(التقاط جرثومة..)': ['Nasocomial Infection Problem'],
    'المعتقدات والمبادئ': ['Respect for beliefs'],
    'اعتراض حول آلية الزيارة': ['Visiting Process'],
    'تأخر النقل (من وإلى..)': ['Delay Transfer(room..)'],
    'خطأ تمريضي(شكة المصل..)': ['IV Problem(IV Insertion Error..)'],
    'مشاكل تتعلق بالإقامة(أجهزة غير كافية,أسرة,تهوئة..)': ['Accomodation Problems(Devices,Beds..)', 'accomodation Problems(area problem..)'],
    'خلل تنسيق مع الأقسام الأخرى': ['Coordination Failure(Team,Other Departements..)'],
    'التواصل الغائب': ['Absent Communication'],
    'التواصل غير الصحيح': ['Incorrect Communication'],
    'ضجة الموظفين': ['Noise(Employee..)'],
    'إنتظار الإجراءات الطبية(صور,فحوصات..)': ['Delay Procedure(Waiting for Imaging,LAB.Tests..)'],
    'تأجيل/تأخير(عملية,تمييل..)': ['Surgical Procedures Delayed'],
    'إجراءات معقّدة/موافقات/تكاليف': ['Complex Procedures/Approvals/Costs'],
    'تأخر تقارير الصور': ['Delayed Procedure(Imaging Reports..)'],
    'عدم الرد على الجرس(غير موصل)': ['Failure to Respond(Nurse call unfunctional)'],
    'التواصل المتأخر': ['Delayed Communication'],
    'خطأ دواء': ['Error -Medication'],
    'تأخر في الرد على الNurse call': ['Delayed Nurse call'],
    'خطأ تمريضي(يهدد سلامة المريض)': ['Technical skills of Staff(that compromise Safety)'],
    'إهمال عام(الرعاية الشخصية,الرعاية الصحية,بيئة آمنة,الدعم النفسي..)': ['Neglect -General (Basic Care,Medical Care,Safe Environment,Physiological Support..)'],
    'عدم الإحترام': ['Disrespect'],
    'الإنتظار للمعاينة': ['Delay Access(Waiting for Consultation..)'],
    'طلبات الدرجة الأولى-': ['First Class Services'],
    'تأخير عام': ['Delay -General(Delay to Respond..)'],
    'خلل في العناية التمريضية(الميل..)': ['Problem Procedure(Foley..)'],
    'أجهزة ولوازم (قسم العمليات,قسم الإيكو..)': ['Equipement & Supplies Problems(OR,Echo..)'],
    'الضجة(أجهوة,أبواب..)': ['Noise(Devices,Doors..)'],
    'العقر السريري': ['Bed Sore Problems'],
    'مضاعفات(اختلاط جراحي..)': ['Complications(Surgical Complication)'],
    'عدم توفر سرير(عادي,عناية..)': ['Bed Unavailablity'],
    'عدم تنسيق حالات المرضى': ['Patient Cases not Organized'],
    'رفض الإستماع المريض/المرافق': ['Dimissing Patients', 'Ignoring Patients'],
    'تأخر إجراء طبي': ['Delay Medical Pocedure'],
    'خلل تنسيق إداري(مع الأقسام الأخرى..)': ['Coordination Problem(Team,Other Departements..)'],
    'عدم التوثيق(اللوازم الطبية,أمر المغادرة..)': ['Documentation Problem(Devices,Discharge..)'],
    'عدم الموافقة(الخطة العلاجية,قرار المغادرة..)': ['Failure to Agree(Treatment Plan,Discharge Decision..)'],
    'الزيارة اليومية للطبيب': ['Daily Doctor Visits(Attending Physician,Consulting Physician)'],
    'إعطاء مواعيد صور متأخرة': ['Delay Access(Imaging Appointment)'],
    'خلل تنسيق طبي(الأطباء,التمريض..)': ['Teamwork Problem(Doctors,Nursing..)'],
    'تحويل المريض من الطوارئ إلى العيادات': ['Disagreement Protocol(ER..)'],
    'خلل في التواصل(إعطاء المعلومات..)': ['Failure to Provide(Information,Treatment..)'],
    'خطأ في إجراء الصورة': ['Imaging Procedure Error'],
    'خلل تنسيق(مع الأقسام الأخرى..)': ['Coordination Problem(Team,Other Departements..)'],
    'خطأ فنيّين(المختبر,الأشعة..)': ['Error Procedure(Lab,X-Ray..)'],
    'عدم الرد على الإتصالات الخارجية': ['Phone Calls Not Anwered'],
    'تأخر حضور الطبيب': ['Delay Procedure(Medical Attendance..)'],
    'تأخر نتائج الفحوصات': ['Delayed Test Results'],
    'خلل في تحديد المواعيد (الصور..)': ['Scheduling Error'],
    'الضجة الصادرة(عدد الزوار,المرضى..)': ['Noise(Accompagnant,Patients..)'],
    'خطأ في التشخيص': ['Error - Diagnosis'],
    'نقص مهارة فنيّ المختبر(نتائج,سحب الدم..)': ['Technician Skills Deficiency(Tests..)'],
    'آلية نقل العيّنة إلى مختبر خارجي': ['Disagreement Protocol(Lab..)'],
    'تقصير في متابعة حالة المريض': ['Error in Monitoring'],
    'عدم تقبّل حضور الطبيب المساعد': ['Failure to Provide(Assistant Visit Issue..)'],
    'مشكلة في الأمن(حدوث سرقة..)': ['Security Problem(Lost ..)'],
    'مشاكل في الأجهزة(المعلوماتية..)': ['IT Problems'],
    'مشاكل في النظافة(الغرفة,الحمام..)': ['Hygene Problem(Room..)'],
    'بطء تنظيف طارىء': ['Delay Cleaning'],
    'تغذية متعدد(الأكل بارد,غير كاف..)': ['Nutritional Problem(Cold Food,Insufficient..)'],
    'بيئة غير آمنة': ['Unsafe Environment'],
    'تغذية(تغليف غير آمن)': ['Unsafe Packaging for Food'],
    'مشاكل تتعلق بالإقامة(محارم,سلة النفايات..)': ['Accomodation Problems(Tissue,Basket...)'],
    'مشاكل تتعلق بالإقامة(جغرافية المكان..)': ['accomodation Problems(area problem..)'],
    'مشاكل تتعلق بالإقامة(التكييف,التدفئة..)': ['Accomodation Problems(Air Conditioning..)'],
    'الضجة من الورشة': ['Noise(Workshop..)'],
    'الضجة': ['Noise(Devices,Doors..)'],
    'ضجة الورشة': ['Noise(Workshop..)'],
    'تأخر إنجاز ملف المغادرة': ['Discharge Delay Problem'],
    'تنسيق حالات المرضى(عدم القدرة على الراحة والنوم..)': ['Patient Case Coordination'],
}


def _lookup_fuzzy(maps: Dict, category: str, value: Optional[str]) -> Tuple[Optional[int], Optional[str]]:
    """
    Two-tier match: exact first (via _lookup), then a fallback against the
    whole bank of known variants per DB value (canonical name + any curated
    LOOKUP_ALIASES entries) -- not just the one canonical string, since some
    real variants don't resemble it closely enough for similarity scoring
    alone to find (see LOOKUP_ALIASES' docstring). Each candidate variant is
    checked via substring containment first (near-certain, score 0.99), then
    difflib.SequenceMatcher ratio as a fallback for spacing/typo noise.

    Returns (matched_id, warning) -- warning is set exactly when the match
    wasn't exact, so callers can flag the row for review (see
    _validate_group's has_fuzzy_match) instead of treating it as clean.
    """
    exact_id = _lookup(maps, category, value)
    if exact_id is not None or not value:
        return exact_id, None

    norm = value.lower().strip()
    candidates = maps.get(category, {})  # {lowercased db name: id}
    aliases = LOOKUP_ALIASES.get(category, {})
    if not candidates:
        return None, None

    best_db_name, best_id, best_score = None, None, 0.0
    for db_name, db_id in candidates.items():
        for variant in (db_name, *aliases.get(db_name, [])):
            score = 0.99 if (norm in variant or variant in norm) else SequenceMatcher(None, norm, variant).ratio()
            if score > best_score:
                best_db_name, best_id, best_score = db_name, db_id, score

    if best_id is not None and best_score >= FUZZY_MATCH_THRESHOLD:
        return best_id, f"'{value}' matched to '{best_db_name.title()}' ({best_score:.0%} match) — please verify"
    return None, None


CLASSIFICATION_SINGLE_CHANNEL_OVERRIDE = 0.90  # a lone channel this strong wins outright, see _match_classification


def _match_classification(maps: Dict, arabic_value: Optional[str], english_value: Optional[str]) -> Tuple[Optional[int], Optional[str]]:
    """
    Dual-channel classification match. The DB's `classifications` lookup is
    Arabic-only, but the source data pairs an Arabic AND an English
    classification per row for the same underlying value -- so rather than
    picking one language and discarding the other's evidence (the original,
    simpler design), this scores every DB candidate on both channels and
    combines them into one decision metric ("the error from both"):

        combined(d) = average of whichever channel(s) actually have input
                      (a blank field is simply not consulted, rather than
                      dragging the average down as a phantom zero)

    A sufficiently strong SINGLE channel (>= CLASSIFICATION_SINGLE_CHANNEL_OVERRIDE)
    wins outright regardless of combined(d) -- otherwise a near-certain match
    on one channel could get wrongly rejected just because the other
    channel's score (e.g. a thin/missing CLASSIFICATION_EN_ALIASES entry)
    drags the average below FUZZY_MATCH_THRESHOLD.

    Returns (matched_id, warning) -- warning is set exactly when the match
    wasn't a clean exact Arabic match, same contract as _lookup_fuzzy.
    """
    exact_id = _lookup(maps, "classifications", arabic_value)
    if exact_id is not None:
        return exact_id, None

    candidates = maps.get("classifications", {})  # {lowercased arabic db name: id}
    ar_norm = (arabic_value or "").lower().strip()
    en_norm = (english_value or "").lower().strip()
    if not candidates or not (ar_norm or en_norm):
        return None, None

    def score(norm: str, variant: str) -> float:
        if not norm or not variant:
            return 0.0
        return 0.99 if (norm in variant or variant in norm) else SequenceMatcher(None, norm, variant).ratio()

    best_db_name, best_id = None, None
    best_combined, best_ar, best_en = -1.0, 0.0, 0.0

    for db_name, db_id in candidates.items():
        ar_score = score(ar_norm, db_name)
        en_score = max((score(en_norm, v.lower().strip()) for v in CLASSIFICATION_EN_ALIASES.get(db_name, [])), default=0.0)

        channels = [s for s, present in ((ar_score, bool(ar_norm)), (en_score, bool(en_norm))) if present]
        combined = sum(channels) / len(channels) if channels else 0.0

        if combined > best_combined:
            best_db_name, best_id = db_name, db_id
            best_combined, best_ar, best_en = combined, ar_score, en_score

    if best_id is None:
        return None, None

    if best_combined >= FUZZY_MATCH_THRESHOLD or max(best_ar, best_en) >= CLASSIFICATION_SINGLE_CHANNEL_OVERRIDE:
        detail = (f"Arabic {best_ar:.0%}" if ar_norm else "") + (", " if ar_norm and en_norm else "") + (f"English {best_en:.0%}" if en_norm else "")
        return best_id, f"Classification matched to '{best_db_name}' ({detail}) — please verify"
    return None, None


def _parse_date_cell(raw: Any, field_label: str) -> Tuple[Optional[date], Optional[str]]:
    """Parse a date-typed template cell, returning (parsed_date, error_message)."""
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None, f"{field_label} is missing"
    if isinstance(raw, datetime):
        return raw.date(), None
    if isinstance(raw, date):
        return raw, None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(raw).strip().split(" ")[0], fmt).date(), None
        except ValueError:
            continue
    return None, f"{field_label} '{raw}' is not a valid date (use YYYY-MM-DD)"


def _validate_group(group_key: str, rows: List[Dict], maps: Dict) -> Dict[str, Any]:
    """
    Validates every row in a group -- every row is checked regardless of
    earlier rows' errors (unlike the old version, which stopped at the
    first bad row), so the review screen can show and let the user fix
    each broken row individually. A group is valid only if every one of
    its rows ends up with zero errors.
    """
    warnings: List[Dict] = []
    row_results: List[Dict[str, Any]] = []
    patient_name = None
    has_fuzzy_match = False

    for row in rows:
        row_number = row.get(ROW_NUMBER_KEY)
        errors: List[Dict[str, str]] = []

        def add_error(field: str, message: str) -> None:
            errors.append({"field": field, "message": message})

        def note_fuzzy(warning: Optional[str]) -> None:
            """Record a _lookup_fuzzy() non-exact-match warning and flag the
            group for review (see has_fuzzy_match in the return dict below)."""
            nonlocal has_fuzzy_match
            if warning:
                warnings.append({"group_key": group_key, "row_number": row_number, "message": warning})
                has_fuzzy_match = True

        # --- Incident Number (Old System) -- the grouping key itself.
        # Blank rows already end up lumped into one MISSING_GROUP_KEY_LABEL
        # bucket by _group_rows(), which would otherwise only surface as a
        # confusing patient-mismatch error once several unrelated blank-key
        # rows collide there. An explicit check gives a direct, actionable
        # message instead, consistent with every other mandatory field. ---
        if not _get(row, COL_INCIDENT_KEY):
            add_error("Incident Number", "Incident Number (Old System) is missing")

        # --- Patient name (consistent across group) ---
        pname = _get(row, COL_PATIENT)
        if not pname:
            add_error("Patient Name", "Patient Name is missing")
        else:
            if patient_name is None:
                patient_name = pname
            elif pname.lower().strip() != patient_name.lower().strip():
                add_error("Patient Name", f"Patient Name '{pname}' differs from group patient '{patient_name}'")

        # --- Incident Date / Received Date ---
        # Both are Excel date-typed, so openpyxl usually hands back a real
        # datetime/date object -- handled directly rather than round-tripped
        # through a string. Plain text is still accepted as a fallback for
        # legacy sheets copy-pasted from before these columns had a date type.
        incident_date_raw = row.get(TEMPLATE_COLUMNS[COL_INCIDENT_DATE][0])
        parsed_incident_date, incident_date_error = _parse_date_cell(incident_date_raw, "Incident Date")
        if incident_date_error:
            add_error("Incident Date", incident_date_error)

        date_raw = row.get(TEMPLATE_COLUMNS[COL_DATE][0])
        parsed_date, received_date_error = _parse_date_cell(date_raw, "Received Date")
        if received_date_error:
            add_error("Received Date", received_date_error)

        # You can't receive a complaint before the incident that prompted it
        # actually happened -- same rule case_service.create_case() enforces
        # for manual entry, checked here too so it shows as a clear preview
        # rejection instead of a raw error at confirm time.
        if parsed_incident_date and parsed_date and parsed_incident_date > parsed_date:
            add_error("Incident Date", "Incident Date cannot be after Received Date")

        # --- Complaint Text ---
        complaint = _get(row, COL_COMPLAINT)
        if not complaint:
            add_error("Complaint Text", "Complaint Text is missing")

        # --- Source (mandatory) ---
        source_name = _get(row, COL_SOURCE)
        source_id, source_warning = _lookup_fuzzy(maps, "sources", source_name)
        note_fuzzy(source_warning)
        if not source_name:
            add_error("Source", "Source is missing")
        elif source_id is None:
            add_error("Source", f"Source '{source_name}' not found in database")

        # --- Issuing Dept (mandatory) ---
        issuing_name = _get(row, COL_ISSUING_DEPT)
        issuing_id, issuing_warning = _lookup_fuzzy(maps, "org_units", issuing_name)
        note_fuzzy(issuing_warning)
        if not issuing_name:
            add_error("Issuing Dept", "Issuing Dept is missing")
        elif issuing_id is None:
            add_error("Issuing Dept", f"Issuing Dept '{issuing_name}' not found — reject")

        # --- Record Type (optional, defaults to Complaint) ---
        record_type_name = _get(row, COL_RECORD_TYPE)
        if record_type_name:
            record_type_id = RECORD_TYPE_IDS.get(record_type_name.lower().strip())
            if record_type_id is None:
                add_error("Record Type", f"Record Type '{record_type_name}' not recognized — use Complaint or Notice")
        else:
            record_type_id = DEFAULT_RECORD_TYPE_ID

        # --- Classification (mandatory for Complaint rows; Notice rows
        # don't require it, mirroring case_service's own Notice exception —
        # drives domain/category/subcategory when present). Dual-channel:
        # the DB's classifications lookup is Arabic-only, but the source
        # pairs an Arabic + English classification per row for the same
        # value, so both are used as independent evidence (see
        # _match_classification / CLASSIFICATION_EN_ALIASES) instead of
        # discarding whichever one isn't the DB's language. ---
        class_name_ar = _get(row, COL_CLASS_AR)
        class_name_en = _get(row, COL_CLASS_EN)
        class_id, class_warning = _match_classification(maps, class_name_ar, class_name_en)
        note_fuzzy(class_warning)
        domain_id = category_id = subcategory_id = None
        if not class_name_ar and not class_name_en:
            if record_type_id != RECORD_TYPE_IDS["notice"]:
                add_error("Classification", "Classification is missing")
        elif class_id is None:
            add_error("Classification", f"Classification '{class_name_ar or class_name_en}' not found — reject")
        else:
            chain = maps.get("classification_chains", {}).get(class_id, {})
            domain_id = chain.get("domain_id")
            category_id = chain.get("category_id")
            subcategory_id = chain.get("subcategory_id")

        # Display-only breadcrumb, computed whenever Classification resolved
        # -- independent of `errors`/`row_data` below, which stay gated on
        # the row being fully valid. An otherwise-invalid row (e.g. a bad
        # Severity) still has a real, resolved Classification chain worth
        # showing the reviewer, not just a blank "derived" placeholder.
        derived_hierarchy = {
            "domain": maps.get("domain_names", {}).get(domain_id),
            "category": maps.get("category_names", {}).get(category_id),
            "subcategory": maps.get("subcategory_names", {}).get(subcategory_id),
        }

        # Display-only resolved values for fields where the raw Excel text
        # doesn't directly match what the review grid needs to highlight --
        # same "independent of errors/row_data" reasoning as derived_hierarchy
        # above. Record Type in particular: the raw cell can be a recognized
        # alias (e.g. "فرصة تحسين" -> Complaint) that the review UI's
        # Complaint/Notice toggle can't detect from the raw text alone, since
        # it isn't literally "Complaint" or "Notice".
        resolved_display = {
            "record_type": (
                "Complaint" if record_type_id == RECORD_TYPE_IDS["complaint"]
                else "Notice" if record_type_id == RECORD_TYPE_IDS["notice"]
                else None
            ),
        }

        # --- Severity / Stage / Harm Level / Feedback Risk Type / Building
        # (all mandatory now -- was warn-and-continue, per explicit decision
        # to require every field with no exceptions) ---
        severity_name = _get(row, COL_SEVERITY)
        severity_id, severity_warning = _lookup_fuzzy(maps, "severities", severity_name)
        note_fuzzy(severity_warning)
        if not severity_name:
            add_error("Severity", "Severity is missing")
        elif severity_id is None:
            add_error("Severity", f"Severity '{severity_name}' not found — reject")

        stage_name = _get(row, COL_STAGE)
        stage_id, stage_warning = _lookup_fuzzy(maps, "stages", stage_name)
        note_fuzzy(stage_warning)
        if not stage_name:
            add_error("Stage", "Stage is missing")
        elif stage_id is None:
            add_error("Stage", f"Stage '{stage_name}' not found — reject")

        harm_name = _get(row, COL_HARM)
        harm_id, harm_warning = _lookup_fuzzy(maps, "harm_levels", harm_name)
        note_fuzzy(harm_warning)
        if not harm_name:
            add_error("Harm Level", "Harm Level is missing")
        elif harm_id is None:
            add_error("Harm Level", f"Harm Level '{harm_name}' not found — reject")

        risk_name = _get(row, COL_RISK)
        risk_id, risk_warning = _lookup_fuzzy(maps, "risk_types", risk_name)
        note_fuzzy(risk_warning)
        if not risk_name:
            add_error("Feedback Risk Type", "Feedback Risk Type is missing")
        elif risk_id is None:
            add_error("Feedback Risk Type", f"Risk Type '{risk_name}' not found — reject")

        building_name = _get(row, COL_BUILDING)
        building_id, building_warning = _lookup_fuzzy(maps, "buildings", building_name)
        note_fuzzy(building_warning)
        if not building_name:
            add_error("Building", "Building is missing")
        elif building_id is None:
            add_error("Building", f"Building '{building_name}' not found — reject")

        # --- Is Inpatient ---
        inpatient_val = _get(row, COL_INPATIENT)
        is_inpatient = str(inpatient_val).strip().lower() in ("yes", "نعم", "1", "true") if inpatient_val else False

        # --- Target Department (mandatory) ---
        # One column, not three: a row represents a case, not an incident,
        # and a case has exactly one target department.
        target_dept_ids: List[int] = []
        dept_name = _get(row, COL_TARGET_DEPT)
        if not dept_name:
            add_error("Target Dept", "Target Dept is missing")
        else:
            dept_id, dept_warning = _lookup_fuzzy(maps, "org_units", dept_name)
            if dept_id is None:
                add_error("Target Dept", f"Target Department '{dept_name}' not found — reject")
            else:
                note_fuzzy(dept_warning)
                target_dept_ids.append(dept_id)

        # --- Doctors: all three slots optional -- warn, don't block, if
        # given but not found; silent if blank. ---
        doctor_ids: List[Tuple[int, str]] = []
        for doc_col in (COL_DOCTOR1, COL_DOCTOR2, COL_DOCTOR3):
            doc_name = _get(row, doc_col)
            if doc_name:
                doc_id, doc_warning = _lookup_fuzzy(maps, "doctors", doc_name)
                if doc_id is None:
                    warnings.append({"group_key": group_key, "row_number": row_number, "message": f"Doctor '{doc_name}' not found — imported without doctor linkage"})
                else:
                    note_fuzzy(doc_warning)
                    doctor_ids.append((doc_id, doc_name))

        # --- Workers: same as doctors, all three slots optional. ---
        worker_ids: List[Tuple[Any, str]] = []
        for wrk_col in (COL_WORKER1, COL_WORKER2, COL_WORKER3):
            wrk_name = _get(row, wrk_col)
            if wrk_name:
                wrk_id, wrk_warning = _lookup_fuzzy(maps, "workers", wrk_name)
                if wrk_id is None:
                    warnings.append({"group_key": group_key, "row_number": row_number, "message": f"Worker '{wrk_name}' not found — skipped"})
                else:
                    note_fuzzy(wrk_warning)
                    worker_ids.append((wrk_id, wrk_name))

        row_data = None
        if not errors:
            row_data = {
                "patient_name": pname,
                "feedback_date": parsed_date,
                "incident_date": parsed_incident_date,
                "record_type_id": record_type_id,
                "feedback_intent_id": _feedback_intent_type_id_for_record_type(record_type_id),
                "source_id": source_id,
                "issuing_org_unit_id": issuing_id,
                "domain_id": domain_id,
                "category_id": category_id,
                "subcategory_id": subcategory_id,
                "classification_id": class_id,
                "severity_id": severity_id,
                "stage_id": stage_id,
                "harm_id": harm_id,
                "risk_type_id": risk_id,
                "building_id": building_id,
                "is_inpatient": is_inpatient,
                "complaint_text": complaint,
                "immediate_action": _get(row, COL_IMMEDIATE),
                "taken_action": _get(row, COL_TAKEN),
                "target_dept_ids": target_dept_ids,
                "doctor_ids": doctor_ids,
                "worker_ids": worker_ids,
            }

        row_results.append({
            "row_number": row_number,
            "errors": errors,
            "data": row_data,
            "raw": {k: v for k, v in row.items() if k != ROW_NUMBER_KEY},
            "derived_hierarchy": derived_hierarchy,
            "resolved_display": resolved_display,
        })

    # --- Patient ambiguity check (group-level, once per group) ---
    # Only worth the directory-search round trip if every row is otherwise
    # clean -- no point checking a group that's already going to be
    # rejected for unrelated reasons. If ambiguous, every row's blocked on
    # it, since it's the shared patient the whole group was validated against.
    is_new_patient = False
    if patient_name and all(not r["errors"] for r in row_results):
        match_count = _count_patient_matches(patient_name)
        if match_count > 1:
            ambiguous_msg = {"field": "Patient Name", "message": f"Patient '{patient_name}' matches {match_count} records — ambiguous, cannot import"}
            for r in row_results:
                r["errors"].append(ambiguous_msg)
                r["data"] = None
        is_new_patient = (match_count == 0)

    return {
        "group_key": group_key,
        "valid": all(not r["errors"] for r in row_results),
        "rows": row_results,
        "warnings": warnings,
        "patient_name": patient_name,
        "is_new_patient": is_new_patient,
        "has_fuzzy_match": has_fuzzy_match,
    }


# ============================================================
# INTERNAL — IMPORT
# ============================================================

def _build_case_service_payload(vrow: Dict, patient_name: str) -> Dict[str, Any]:
    """
    Translate an import_service validated-row dict into the data shape
    case_service.create_case() expects. Incident Date and Received Date are
    now distinct template columns (both mandatory, validated against each
    other in _validate_group). requires_explanation defaults to False — the
    FSM's own red-flag/never-event check, driven by the template's optional
    Risk Type column, is what actually opens the explanation workflow for
    imported rows that need it (moot anyway under save_mode='import_closed',
    which keeps every imported case Closed regardless).
    """
    return {
        "record_type_id": vrow.get("record_type_id", DEFAULT_RECORD_TYPE_ID),
        "patient_name": patient_name,
        "feedback_received_date": vrow["feedback_date"],
        "incident_date": vrow["incident_date"],
        "feedback_intent_type_id": vrow["feedback_intent_id"],
        "source_id": vrow.get("source_id"),
        "issuing_department_id": vrow["issuing_org_unit_id"],
        "domain_id": vrow.get("domain_id"),
        "category_id": vrow.get("category_id"),
        "subcategory_id": vrow.get("subcategory_id"),
        "classification_id": vrow.get("classification_id"),
        "severity_id": vrow.get("severity_id"),
        "stage_id": vrow.get("stage_id"),
        "harm_id": vrow.get("harm_id"),
        "clinical_risk_type_id": vrow.get("risk_type_id") or DEFAULT_CLINICAL_RISK_TYPE_ID,
        "requires_explanation": False,
        "building_id": vrow.get("building_id") or 1,  # default building when blank
        "is_inpatient": vrow.get("is_inpatient", False),
        "is_morbidity": False,
        "complaint_text": vrow["complaint_text"],
        "immediate_action": vrow.get("immediate_action") or "",
        "taken_action": vrow.get("taken_action") or "",
        "target_department_ids": vrow.get("target_dept_ids") or [],
        "doctors": [
            {"doctor_id": doc_id, "doctor_name": doc_name}
            for doc_id, doc_name in vrow.get("doctor_ids", [])
        ],
        "employees": [
            {"employee_id": emp_id, "full_name": emp_name}
            for emp_id, emp_name in vrow.get("worker_ids", [])
        ],
    }


def _import_group(
    group_key: str,
    validated_rows: List[Dict],
    is_new_patient: bool,
    created_by_user_id: int,
    import_batch_id: Optional[int] = None,
) -> Dict[str, Any]:
    """
    Create the shared incident parent for this group, then create each row's
    case through the centralized case_service.create_case() (context=
    'BulkImport') — the same validation/FSM/ML-job-registration logic manual
    insert uses, rather than a separate raw-SQL implementation.

    Each row is its own atomic unit (case + its ml.EmbeddingProcessingJob
    commit together, inside create_case()). This means a later row failing
    within a group does NOT roll back earlier rows already committed in the
    same group — a deliberate trade-off (see ML_ARCHITECTURE_DECISION_RECORD.md
    Stage 4) that favors per-case durability and ML-job guarantees over
    whole-group atomicity. Partial success is reported accurately via
    rows_imported/rows_failed/row_errors rather than silently rolling back
    already-valid rows or misreporting the whole group as rejected.
    """
    first = validated_rows[0]
    patient_name = first["patient_name"]

    conn = get_connection()
    conn.autocommit = False
    cursor = conn.cursor()

    try:
        # Create reserve patient if new
        if is_new_patient:
            import_db.create_reserve_patient(patient_name, created_by_user_id, cursor)

        # Create the shared incident parent for this group
        incident_id = import_db.insert_incident(
            patient_name=patient_name,
            feedback_intent_type_id=first["feedback_intent_id"],
            issuing_org_unit_id=first["issuing_org_unit_id"],
            building_id=first.get("building_id") or 1,
            is_inpatient=first.get("is_inpatient", False),
            created_by_user_id=created_by_user_id,
            cursor=cursor,
        )
        conn.commit()
    except Exception:
        conn.rollback()
        cursor.close()
        conn.close()
        raise
    cursor.close()
    conn.close()

    rows_imported = 0
    row_errors: List[Dict[str, Any]] = []

    for idx, vrow in enumerate(validated_rows, start=1):
        try:
            data = _build_case_service_payload(vrow, patient_name)
            result = create_case(data, context='BulkImport', save_mode='import_closed')

            if not result.get("success"):
                row_errors.append({"row_index": idx, "error": result.get("message", "Unknown error")})
                continue

            case_id = int(result["id"])

            # Link this case to the group's shared incident parent
            assign_case_to_incident(case_id, incident_id)

            # Traceability mapping (case <-> source row), not a duplicate
            # guard -- see the note in _validate_and_group on why Incident
            # Group Key can't be a cross-upload identifier. Scoped by
            # import_batch_id (always a real, unique batch at this point)
            # so ml.ImportSourceRecordMap's UNIQUE(ExternalSourceSystem,
            # ExternalRecordID) constraint can't collide between two
            # different uploads that happen to reuse the same group key.
            external_record_id = f"{import_batch_id}:{group_key}#{idx}"
            map_conn = get_connection()
            map_cursor = map_conn.cursor()
            try:
                ml_import_batch_db.record_source_map(
                    map_cursor, import_batch_id, EXCEL_IMPORT_SOURCE_SYSTEM, external_record_id, case_id
                )
                map_conn.commit()
            finally:
                map_cursor.close()
                map_conn.close()

            rows_imported += 1

        except Exception as exc:
            row_errors.append({"row_index": idx, "error": str(exc)})

    return {
        "group_key": group_key,
        "incident_id": incident_id,
        "rows_imported": rows_imported,
        "rows_failed": len(row_errors),
        "row_errors": row_errors,
        "new_patient": is_new_patient,
    }


# ============================================================
# INTERNAL — REJECTED EXCEL
# ============================================================

def _generate_import_report_excel(
    imported_with_rows: List[Tuple[Dict[str, Any], List[Dict]]],
    rejected_groups: List[Dict[str, Any]],
) -> BytesIO:
    """
    Full receipt of a confirmed batch: every original row, green if imported
    (with its new system Incident Number) or red/blue if rejected (with the
    reason) -- matches the original ask (green=imported+new ID, red=failure
    +reason), extended with a distinct color for duplicates so that reason
    isn't confused with a plain validation failure.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Report"
    ws.sheet_view.rightToLeft = True

    headers = [col[0] for col in TEMPLATE_COLUMNS] + ["Result", "System ID / Reason"]
    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    duplicate_fill = PatternFill("solid", fgColor="BDD7EE")

    row_i = 2
    for imp_result, group_rows in imported_with_rows:
        incident_number = f"INC-{imp_result['incident_id']:06d}"
        for row_data in group_rows:
            for col_i, (header, _, _, _) in enumerate(TEMPLATE_COLUMNS, start=1):
                cell = ws.cell(row=row_i, column=col_i, value=row_data.get(header))
                cell.fill = green_fill
            ws.cell(row=row_i, column=len(TEMPLATE_COLUMNS) + 1, value="Imported").fill = green_fill
            ws.cell(row=row_i, column=len(TEMPLATE_COLUMNS) + 2, value=incident_number).fill = green_fill
            row_i += 1

    for group in rejected_groups:
        fill = duplicate_fill if group.get("status") == "duplicate" else red_fill
        result_label = "Duplicate" if group.get("status") == "duplicate" else "Rejected"
        for row_data in group["rows"]:
            for col_i, (header, _, _, _) in enumerate(TEMPLATE_COLUMNS, start=1):
                cell = ws.cell(row=row_i, column=col_i, value=row_data.get(header))
                cell.fill = fill
            ws.cell(row=row_i, column=len(TEMPLATE_COLUMNS) + 1, value=result_label).fill = fill
            ws.cell(row=row_i, column=len(TEMPLATE_COLUMNS) + 2, value=group["reason"]).fill = fill
            row_i += 1

    for col_i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 20
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _empty_report(message: str) -> Dict[str, Any]:
    return {
        "summary": {
            "total_groups": 0, "imported_groups": 0, "rejected_groups": 0,
            "total_rows": 0, "imported_rows": 0, "rejected_rows": 0,
            "new_patients_created": 0, "warnings_count": 0,
        },
        "imported": [],
        "rejected": [],
        "warnings": [{"group_key": "-", "message": message}],
        "rejected_excel_b64": None,
    }


def list_import_batches(limit: int = 50) -> List[Dict[str, Any]]:
    """Batch history for the review UI — mirrors the ML Training Run Artifacts table layout."""
    conn = get_connection()
    cursor = conn.cursor()
    try:
        batches = ml_import_batch_db.list_batches(cursor, limit=limit)
    finally:
        cursor.close()
        conn.close()

    for b in batches:
        for key in ("UploadedAt", "CompletedAt"):
            if b.get(key) is not None:
                b[key] = b[key].isoformat()
    return batches
